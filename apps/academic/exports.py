import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from django.db import connection
from apps.academic.models import StudentGrade
from apps.students.models import Enrollment

class ExportEngine:
    """
    Motor de Exportação Enterprise SOTARQ.
    Consolida Listas de Chamada e Pautas de Aproveitamento.
    """

    @staticmethod
    def generate_class_excel(class_obj):
        """Gera Lista de Frequência em Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Lista - {class_obj.name}"
        tenant = connection.tenant

        # Estilização
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        # Cabeçalho
        ws.merge_cells('A1:E1')
        ws['A1'] = f"{tenant.name.upper()} - ANO LECTIVO {class_obj.academic_year.name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = center_align

        headers = ['Nº', 'Nº MATRÍCULA', 'NOME COMPLETO DO ESTUDANTE', 'GÉNERO', 'OBSERVAÇÕES']
        ws.append([]) 
        ws.append(headers)
        
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Dados (Ordem alfabética oficial)
        enrollments = Enrollment.objects.filter(
            class_room=class_obj, status='active'
        ).select_related('student').order_by('student__full_name')

        for idx, enr in enumerate(enrollments, 1):
            ws.append([idx, enr.student.registration_number, enr.student.full_name.upper(), enr.student.gender, ""])

        ws.column_dimensions['C'].width = 50
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def generate_final_pauta_excel(klass):
        """Gera a Pauta Final Consolidada (I, II e III Trimestres) - Decreto 424/25."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PAUTA FINAL"
        tenant = connection.tenant

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        subjects = klass.grade_level.subjects.all().order_by('name')
        
        ws.cell(row=1, column=1, value=f"{tenant.name.upper()} - PAUTA FINAL - {klass.academic_year.name}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(subjects)*4 + 3)

        headers = ["Nº", "Nome Completo"]
        for s in subjects:
            headers.extend([f"{s.name[:15]} (T1)", "(T2)", "(T3)", "(MF)"])
        headers.append("RESULTADO FINAL")

        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=text)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center_align

        enrollments = klass.enrollments_records.filter(status='active').select_related('student').order_by('student__full_name')
        
        for row_idx, enrollment in enumerate(enrollments, 4):
            student = enrollment.student
            ws.cell(row=row_idx, column=1, value=row_idx-3)
            ws.cell(row=row_idx, column=2, value=student.full_name.upper())

            col_ptr = 3
            failed_subjects = 0
            
            for subject in subjects:
                grade = student.academic_grades.filter(subject=subject, klass=klass).first()
                if grade:
                    ws.cell(row=row_idx, column=col_ptr, value=grade.mt1)
                    ws.cell(row=row_idx, column=col_ptr+1, value=grade.mt2)
                    ws.cell(row=row_idx, column=col_ptr+2, value=grade.mt3)
                    
                    # MF com arredondamento visual
                    mf_cell = ws.cell(row=row_idx, column=col_ptr+3, value=round(grade.mf))
                    
                    pass_grade = 5 if klass.grade_level.level_index <= 6 else 10
                    if grade.mf < pass_grade:
                        mf_cell.font = Font(color="FF0000", bold=True)
                        failed_subjects += 1
                
                col_ptr += 4

            # Resultado Lógico
            result = "TRANSITA" if failed_subjects <= 2 else "RETIDO"
            if enrollment.status == 'failed': result = "RETIDO (FALTAS)"
            
            res_cell = ws.cell(row=row_idx, column=col_ptr, value=result)
            res_cell.font = Font(bold=True, color="0000FF" if "TRANSITA" in result else "FF0000")

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def generate_class_pdf(class_obj):
        """Gera Lista de Presença em PDF (Landscape)."""
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)
        tenant = connection.tenant

        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(width/2, height - 50, tenant.name.upper())
        p.setFont("Helvetica", 10)
        p.drawCentredString(width/2, height - 70, f"LISTA DE PRESENÇA - TURMA: {class_obj.name} ({class_obj.academic_year.name})")

        y = height - 120
        p.drawString(50, y, "Nº")
        p.drawString(80, y, "NOME DO ESTUDANTE")
        
        # Grelha de assinatura (1-15 aulas)
        for i in range(1, 16):
            p.drawString(300 + (i*30), y, str(i))

        enrollments = Enrollment.objects.filter(class_room=class_obj).select_related('student').order_by('student__full_name')
        y -= 20
        for idx, enr in enumerate(enrollments, 1):
            p.setFont("Helvetica", 9)
            p.drawString(50, y, str(idx))
            p.drawString(80, y, enr.student.full_name[:40].upper())
            p.line(50, y - 5, width - 50, y - 5)
            y -= 18
            if y < 50:
                p.showPage()
                y = height - 50

        p.showPage()
        p.save()
        return buffer.getvalue()

    @staticmethod
    def generate_minipauta_excel(klass, subject, term):
        """Gera Minipauta detalhada para Conselhos de Notas (T1, T2 ou T3)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Minipauta - {subject.name[:10]} - T{term}"
        tenant = connection.tenant

        # 1. Estilização Enterprise
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # 2. Cabeçalho Dinâmico
        ws.merge_cells('A1:F1')
        ws['A1'] = f"{tenant.name.upper()} - MINIPAUTA TRIMESTRAL"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:F2')
        ws['A2'] = f"TURMA: {klass.name} | DISCIPLINA: {subject.name} | TRIMESTRE: {term}º"
        ws['A2'].alignment = center_align

        # 3. Colunas de Dados
        headers = ['Nº', 'NOME COMPLETO DO ESTUDANTE', 'MAC', 'NPP', 'NPT', f'MT {term}']
        ws.append([]) # Linha vazia
        ws.append(headers)

        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # 4. Busca de Dados com Rigor de Ordenação
        grades = StudentGrade.objects.filter(
            klass=klass, subject=subject
        ).select_related('student').order_by('student__full_name')

        for idx, g in enumerate(grades, 1):
            # Extração dinâmica baseada no trimestre (term)
            mac = getattr(g, f'mac{term}', 0)
            npp = getattr(g, f'npp{term}', 0)
            npt = getattr(g, f'npt{term}', 0)
            mt = getattr(g, f'mt{term}', 0)

            row = [idx, g.student.full_name.upper(), float(mac), float(npp), float(npt), float(mt)]
            ws.append(row)
            
            # Formatação de negativas (Menor que 5 no primário, menor que 10 no secundário)
            pass_mark = 5 if klass.grade_level.level_index <= 6 else 10
            if mt < pass_mark:
                ws.cell(row=ws.max_row, column=6).font = Font(color="FF0000", bold=True)

        # Ajustes de Layout
        ws.column_dimensions['B'].width = 45
        for col in ['C', 'D', 'E', 'F']: ws.column_dimensions[col].width = 10

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


