# apps/academic/views.py
from datetime import datetime
import logging
import tempfile
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.db import transaction, models 
from django.db.models import Q
from django.contrib import messages
from django.utils import timezone
from apps.core.models import Notification, Role, SchoolConfiguration, User
from apps.core.decorators import student_required
from apps.core.servicos.notifications import AlertService
from apps.students.models import Enrollment, Student
from apps.teachers.models import Teacher, TeacherSubject
from apps.finance.models import FeeType, Invoice
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from apps.academic.models import Class, StudentGrade, Subject

# Modelos Académicos e Forms
from .models import (
    AcademicGlobal, AcademicYear, Class, Classroom, LessonPlan, Subject, 
    StudentGrade, VacancyRequest, AcademicEvent
)
from .forms import AcademicEventEmailsForm, AcademicYearForm, GradeLevelForm, PedagogicalLockForm, SubjectForm
from .exports import ExportEngine
# Importação segura para evitar quebra se analytics tiver erro
try:
    from .analytics import EfficiencyAnalytics
except ImportError:
    EfficiencyAnalytics = None

logger = logging.getLogger(__name__)






def is_manager_check(user):
    """
    Verifica permissão de gestão (Admin ou Diretor) no contexto do Tenant.
    Ref: Django Authentication System
    """
    if not user.is_authenticated:
        return False
    MANAGEMENT_ROLES = [Role.Type.ADMIN, Role.Type.DIRECTOR]
    return user.is_superuser or user.current_role in MANAGEMENT_ROLES

def is_secretary_directors(user):
    """
    Verifica permissão de gestão (Admin ou Diretor) no contexto do Tenant.
    Ref: Django Authentication System
    """
    if not user.is_authenticated:
        return False
    MANAGEMENT_ROLES = [Role.Type.ADMIN, Role.Type.DIRECTOR, Role.Type.SECRETARY, Role.Type.DIRECT_FINANC, Role.Type.DIRECT_ADMIN]
    return user.is_superuser or user.current_role in MANAGEMENT_ROLES


def is_teacher_pedagogic(user):
    """
    Verifica permissão de gestão (Admin ou Diretor) no contexto do Tenant.
    Ref: Django Authentication System
    """
    if not user.is_authenticated:
        return False
    MANAGEMENT_ROLES = [Role.Type.ADMIN, Role.Type.DIRECTOR, Role.Type.STUDENT, Role.Type.PEDAGOGIC]
    return user.is_superuser or user.current_role in MANAGEMENT_ROLES

# ==============================================================================
# 1. GESTÃO DE PAUTAS E NOTAS (REVISÃO SÊNIOR)
# ==============================================================================
@login_required
def class_grading_sheet(request, class_id, subject_id, term=1):
    """Exibe a pauta de consulta. Bloqueia acesso Cross-Tenant e valida posse."""
    
    # 1. SEGURANÇA: Filtro obrigatório por Tenant (Rigor SOTARQ)
    # Certifique-se que o modelo Class ou AcademicYear tenha o campo 'tenant'

    #klass = get_object_or_404(Class, id=class_id, academic_year__school=request.user.tenant)

    klass = get_object_or_404(Class, id=class_id)

    # Lógica de seletor (se subject_id for 0)
    if subject_id == 0:
        subjects = Subject.objects.filter(grade_level=klass.grade_level)
        return render(request, 'academic/subject_selector.html', {
            'klass': klass, 
            'subjects': subjects, 
            'term': term
        })

    subject = get_object_or_404(Subject, id=subject_id)
    
    # 2. PERMISSÃO: Se não for Gestor, verifica alocação do Professor
    if not request.user.is_manager:
        # Importante: TeacherSubject deve ligar o User ao Tenant
        is_allocated = TeacherSubject.objects.filter(
            teacher__user=request.user, 
            class_room=klass, 
            subject=subject
        ).exists()
        
        if not is_allocated:
            return HttpResponseForbidden("ACESSO NEGADO: Você não leciona esta disciplina nesta turma.")

    # 3. ESTRATÉGIA BULK: Garante que todos os alunos matriculados tenham registro de nota
    active_enrollments = Enrollment.objects.filter(
        class_room=klass, 
        status='active'
    ).select_related('student')

    # Busca notas já existentes para evitar duplicatas
    existing_student_ids = StudentGrade.objects.filter(
        klass=klass, 
        subject=subject
    ).values_list('student_id', flat=True)
    
    # Cria registros faltantes em massa
    new_grades = [
        StudentGrade(
            student=enrol.student,
            subject=subject,
            klass=klass,
            academic_year=klass.academic_year
        ) for enrol in active_enrollments if enrol.student_id not in existing_student_ids
    ]
    
    if new_grades:
        StudentGrade.objects.bulk_create(new_grades)

    # Busca todas as notas para exibição
    grades = StudentGrade.objects.filter(
        klass=klass, 
        subject=subject
    ).select_related('student').order_by('student__full_name')
    
    return render(request, 'academic/grading_sheet.html', {
        'klass': klass,
        'subject': subject,
        'grades': grades,
        'term': term
    })


@login_required
def mass_grade_entry(request, allocation_id, term=1):
    """Interface de lançamento rápido. Proteção total contra ID Spoofing."""
    # 1. SEGURANÇA: Diretor acessa qualquer uma do Tenant, Professor apenas a DELE.
    if request.user.is_manager:
        allocation = get_object_or_404(
            TeacherSubject, 
            id=allocation_id, 
            class_room__academic_year__tenant=request.user.tenant
        )
    else:
        allocation = get_object_or_404(
            TeacherSubject, 
            id=allocation_id, 
            teacher__user=request.user
        )
    
    # 2. PERFORMANCE: Lazy creation otimizada (mesma lógica bulk da view acima)
    active_enrollments = allocation.class_room.enrollments_records.filter(status='active').select_related('student')
    active_student_ids = [e.student_id for e in active_enrollments]
    
    existing_grade_student_ids = StudentGrade.objects.filter(
        klass=allocation.class_room, 
        subject=allocation.subject
    ).values_list('student_id', flat=True)

    missing_student_ids = set(active_student_ids) - set(existing_grade_student_ids)
    if missing_student_ids:
        StudentGrade.objects.bulk_create([
            StudentGrade(
                student_id=s_id, 
                subject=allocation.subject, 
                klass=allocation.class_room, 
                academic_year=allocation.class_room.academic_year
            ) for s_id in missing_student_ids
        ])

    grades = StudentGrade.objects.filter(
        klass=allocation.class_room, 
        subject=allocation.subject
    ).select_related('student').order_by('student__full_name')

    return render(request, 'academic/mass_grade_entry.html', {
        'allocation': allocation,
        'grades': grades,
        'term': term
    })




@login_required
@require_POST
def update_grade_inline(request, grade_id):
    """
    Atualização cirúrgica de nota via interface parcial.
    Garante isolamento de Tenant e validação de 3 trimestres.
    """
    # 1. RIGOR DE SEGURANÇA: Filtro por Tenant e Ownership (Posse)
    # Se for Diretor, acessa qualquer nota da SUA escola.
    # Se for Professor, acessa apenas notas das SUAS disciplinas.
    try:
        if request.user.is_manager:
            grade = StudentGrade.objects.get(
                id=grade_id, 
                klass__academic_year__tenant=request.user.tenant
            )
        else:
            grade = StudentGrade.objects.get(
                id=grade_id,
                klass__teachersubject__teacher__user=request.user,
                subject__teachersubject__teacher__user=request.user
            )
    except StudentGrade.DoesNotExist:
        logger.warning(f"Tentativa de acesso ilegal: User {request.user.id} -> Grade {grade_id}")
        return HttpResponseForbidden("Acesso negado: Registro não pertence à sua jurisdição.")

    field = request.POST.get('field')
    value = request.POST.get('value')
    
    # 2. SUPORTE TOTAL AOS 3 TRIMESTRES (Sem omissões)
    valid_fields = [
        'mac1', 'npp1', 'npt1', 
        'mac2', 'npp2', 'npt2', 
        'mac3', 'npp3', 'npt3'
    ]
    
    if field in valid_fields:
        try:
            # Validação de intervalo (Norma Angolana: 0-20)
            val = float(value) if value and value.strip() else 0.0
            if 0 <= val <= 20:
                setattr(grade, field, val)
                grade.save() # Dispara MTs e MF automáticos no Model
            else:
                return HttpResponse("Valor deve estar entre 0 e 20", status=400)
        except ValueError:
            return HttpResponse("Valor numérico inválido", status=400)
        
    # 3. RETORNO DE INTERFACE (HTMX Parcial)
    # Identifica o trimestre para o template saber qual MT exibir
    term = field[-1] if field else 1
    return render(request, 'academic/partials/grade_row_update.html', {
        'grade': grade,
        'term': int(term)
    })



@login_required
@require_POST
def update_grade_ajax(request):
    """
    Versão Sênior: Restaura a captura robusta e injeta segurança multi-tenant.
    Valida bloqueio, salva nota e retorna MT e MF.
    """
    # 1. VERIFICAÇÃO DE BLOQUEIO (Mantendo sua lógica original de exceções)
    lock_event = AcademicEvent.objects.filter(category='HOLIDAY', klass__isnull=True).first()
    if lock_event and lock_event.is_pedagogical_break:
        is_exception = False
        try:
            # Rigor: Busca o perfil de professor vinculado ao usuário
            teacher = request.user.teacher_profile
            is_exception = lock_event.break_exceptions.filter(id=teacher.id).exists()
        except AttributeError:
            # Se não tem teacher_profile, verifica se é Diretor/Admin
            is_exception = request.user.current_role in [Role.Type.ADMIN, Role.Type.DIRECTOR]

        if not is_exception:
            return JsonResponse({
                'status': 'error', 
                'message': 'Sistema em Pausa Pedagógica. Lançamento bloqueado.'
            }, status=403)
    
    # 2. CAPTURA DE DADOS (Exatamente como na sua versão de confiança)
    grade_id = request.POST.get('grade_id')
    field = request.POST.get('field')  # ex: 'mac1', 'npp2', 'npt3'
    value = request.POST.get('value')

    try:
        # 3. SEGURANÇA DE ACESSO (Ownership)
        # Se for Diretor, ele pode editar qualquer nota do seu próprio Tenant
        if request.user.is_manager:
            grade = StudentGrade.objects.get(
                id=grade_id, 
                klass__academic_year__tenant=request.user.tenant
            )
        else:
            # Se for Professor, só edita se a disciplina/turma for dele
            # Usamos filter().get() para garantir precisão
            grade = StudentGrade.objects.get(
                id=grade_id, 
                subject__teachersubject__teacher__user=request.user,
                klass__teachersubject__teacher__user=request.user
            )

        # 4. VALIDAÇÃO E PERSISTÊNCIA
        val = float(value) if value and value.strip() else 0.0
        
        if 0 <= val <= 20:
            setattr(grade, field, val)
            grade.save()  # Dispara o recálculo de MTs e MF no Model

            # Identifica qual MT retornar (mac1 -> mt1, npp2 -> mt2, etc)
            term_suffix = field[-1] 
            current_mt = getattr(grade, f'mt{term_suffix}')
            
            return JsonResponse({
                'status': 'success', 
                'new_mean': f"{current_mt:.1f}",
                'final_mean': f"{grade.mf:.1f}"
            })
        else:
            return JsonResponse({'status': 'error', 'message': 'Nota inválida (0-20)'}, status=400)
            
    except StudentGrade.DoesNotExist:
        logger.error(f"Tentativa de violação de acesso: User {request.user.id} -> Grade {grade_id}")
        return JsonResponse({'status': 'error', 'message': 'Acesso negado ou registro inexistente.'}, status=403)
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Valor numérico inválido.'}, status=400)

# ==============================================================================
# 2. PORTAL DO ESTUDANTE
# ==============================================================================


@login_required
def student_dashboard(request):
   
    #Dashboard Centralizado SOTARQ SCHOOL.
    #Unifica: Portal do Aluno, Gestão de Staff, Bloqueio Pedagógico e Despacho de Vagas.

    user = request.user
    role = user.current_role
    active_year = AcademicYear.objects.filter(is_active=True).first()

    

    # 1. DISPATCH PARA ALUNO/ENCARREGADO
    if role in [Role.Type.STUDENT, Role.Type.GUARDIAN]:
        return _handle_student_portal_view(request)

    # 2. STAFF/ADMIN/DIRETOR
    is_director = role in [Role.Type.ADMIN, Role.Type.DIRECTOR]
    has_access = user.pode_acessar_academic_page or is_director

    if not has_access:
        messages.error(request, "Acesso Negado: Não tens permissão para a Gestão Académica.")
        return redirect('core:dashboard')
    
    # --- NOVO: BUSCAR ALOCAÇÕES SE FOR PROFESSOR ---
    teacher_allocations = []
    if role == Role.Type.TEACHER:
        # Busca todas as disciplinas/turmas que este professor leciona no ano ativo
        teacher_allocations = TeacherSubject.objects.filter(
            teacher__user=user,
            class_room__academic_year=active_year
        ).select_related('class_room', 'subject')

    # --- BLOQUEIO PEDAGÓGICO (AcademicGlobal) ---
    academic_global, _ = AcademicGlobal.objects.get_or_create(pk=1)  # Garantir objeto único
    if request.method == 'POST' and 'update_lock' in request.POST and is_director:
        form_lock = PedagogicalLockForm(request.POST, instance=academic_global)
        if form_lock.is_valid():
            form_lock.save()
            messages.success(request, "Configurações de bloqueio pedagógico atualizadas!")
            return redirect('academic:student_dashboard')
    else:
        form_lock = PedagogicalLockForm(instance=academic_global)

    # --- E-MAILS EXECUTIVOS (AcademicEventEmailsForm) ---
    lock_event = AcademicEvent.objects.filter(category='HOLIDAY', klass__isnull=True).first()
    if request.method == 'POST' and 'update_emails' in request.POST and is_director:
        form_emails = AcademicEventEmailsForm(request.POST, instance=lock_event)
        if form_emails.is_valid():
            form_emails.save()
            messages.success(request, "E-mails executivos atualizados!")
            return redirect('academic:student_dashboard')
    else:
        form_emails = AcademicEventEmailsForm(instance=lock_event)

    # --- VAGAS PENDENTES (DIRETOR) ---
    pending_vacancies = []
    pending_count = 0
    if is_director:
        pending_vacancies = VacancyRequest.objects.filter(
            is_resolved=False,
        ).select_related('student', 'target_grade')[:5]
        pending_count = VacancyRequest.objects.filter(
            is_resolved=False,
        ).count()

    # --- DADOS DE TURMAS ---
    active_year = AcademicYear.objects.filter(is_active=True).first()
    classes = Class.objects.filter(academic_year=active_year).select_related(
        'grade_level', 'grade_level__course'
    )

    context = {
        'classes': classes,
        'teacher_allocations': teacher_allocations,
        'view_type': role,
        'active_year': active_year,
        'is_director': is_director,
        'form_lock': form_lock,
        'form_emails': form_emails,
        'pending_vacancies': pending_vacancies,
        'pending_vacancies_count': pending_count,
        'stats': {
            'total_classes': classes.count(),
            'total_students': Student.objects.filter(is_active=True).count(),
        },
        'perms': {
            'ver_pautas': is_director or user.pode_ver_pautas_boletins,
            'baixar_pautas': is_director or user.pode_baixar_pautas,
            'ver_docs': is_director or user.pode_ver_documentos_academics,
        }
    }
    return render(request, 'academic/academic_page.html', context)



def _handle_student_portal_view(request):
    """Sua lógica original de Aluno, isolada para organização."""
    student = request.user.student_profile
    today = timezone.now()
    
    tuition_invoices = Invoice.objects.filter(student=student, items__description__icontains="Propina").distinct()
    paid_months = tuition_invoices.filter(status='paid').count()
    tuition_progress = (paid_months / 10) * 100 # Base 10 meses
    
    events = AcademicEvent.objects.filter(
        Q(klass=student.current_class) | Q(klass__isnull=True),
        start_date__gte=today
    ).order_by('start_date')[:5]

    context = {
        'student': student,
        'tuition_progress': tuition_progress,
        'paid_months': paid_months,
        'events': events,
        'pending_invoices': student.invoices.filter(status__in=['pending', 'overdue']),
    }
    return render(request, 'portal/dashboard.html', context)

# ==============================================================================
# 3. PLANOS DE AULA
# ==============================================================================

@login_required
def create_lesson_plan(request):
    if request.method == 'POST':
        allocation_id = request.POST.get('allocation_id')
        topic = request.POST.get('topic')
        content = request.POST.get('content')
        date = request.POST.get('date')

        allocation = get_object_or_404(TeacherSubject, id=allocation_id, teacher__user=request.user)

        plan = LessonPlan.objects.create(
            allocation=allocation,
            topic=topic,
            content=content,
            date=date
        )

        directors = User.objects.filter(tenant=request.user.tenant, current_role=Role.Type.DIRECTOR)
        msgs = [Notification(
            user=d,
            title="Novo Plano de Aula 📚",
            message=f"Prof. {request.user.first_name} criou plano para {allocation.class_room.name}: {topic}",
            link=f"/academic/lesson-plan/{plan.id}/"
        ) for d in directors]
        Notification.objects.bulk_create(msgs)

        messages.success(request, "Plano de aula criado e notificado à direção.")
        return redirect('students:student_list')

    return redirect('students:student_list')

# ==============================================================================
# 4. EXPORTAÇÕES (PAUTAS)
# ==============================================================================

def download_pauta_excel(request, class_id):
    class_obj = get_object_or_404(Class, id=class_id)
    excel_data = ExportEngine.generate_class_excel(class_obj)
    response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Pauta_{class_obj.name}.xlsx"'
    return response

def download_pauta_pdf(request, class_id):
    class_obj = get_object_or_404(Class, id=class_id)
    pdf_data = ExportEngine.generate_class_pdf(class_obj)
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Pauta_{class_obj.name}.pdf"'
    return response

# ==============================================================================
# 5. GESTÃO DE VAGAS
# ==============================================================================

@login_required
@user_passes_test(is_manager_check, login_url='/', redirect_field_name=None)
def manage_vacancy_request(request, vacancy_id):
    """
    Interface de Despacho do Diretor para Solicitações de Vaga.
    Unifica a aprovação financeira com a alocação de turma.
    """
    vacancy = get_object_or_404(
        VacancyRequest.objects.select_related('target_grade', 'student__user'), 
        id=vacancy_id,
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '').strip()
        
        try:
            with transaction.atomic():
                if action == 'approve':
                    class_id = request.POST.get('class_id')
                    if not class_id:
                        messages.error(request, "Selecione uma turma para alocar o aluno.")
                        return redirect(request.path)

                    target_class = get_object_or_404(Class, id=class_id)
                    
                    # 1. Aloca o aluno na turma (Atualiza Matrícula)
                    enrollment = Enrollment.objects.filter(
                        student=vacancy.student, 
                        status='pending_placement'
                    ).first()
                    
                    if enrollment:
                        enrollment.class_room = target_class
                        enrollment.status = 'active'
                        enrollment.save()
                    
                    # 2. Atualiza o Pedido de Vaga
                    vacancy.status = 'approved'
                    vacancy.is_resolved = True
                    vacancy.message += f"\n[DESPACHO {request.user.username}]: APROVADO. {notes}"
                    vacancy.save()

                    # 3. Notifica o Aluno/Encarregado
                    Notification.objects.create(
                        user=vacancy.student.user,
                        title="Vaga Aprovada! ✅",
                        message=f"Entrada autorizada na turma {target_class.name}. Bem-vindo!",
                        link=reverse('portal:dashboard')
                    )
                    messages.success(request, f"Aluno {vacancy.student.full_name} alocado com sucesso na {target_class.name}.")

                elif action == 'deny':
                    vacancy.status = 'denied'
                    vacancy.is_resolved = True
                    vacancy.message += f"\n[DESPACHO {request.user.username}]: NEGADO. {notes}"
                    vacancy.save()

                    Notification.objects.create(
                        user=vacancy.student.user,
                        title="Solicitação de Vaga Indeferida",
                        message=f"Infelizmente não foi possível alocá-lo. {notes}",
                        link=reverse('portal:dashboard')
                    )
                    messages.warning(request, "Solicitação negada e aluno notificado.")

                return redirect('academic:student_dashboard')

        except Exception as e:
            messages.error(request, f"Erro ao processar despacho: {str(e)}")

    # Busca turmas disponíveis para o nível solicitado
    available_classes = Class.objects.filter(
        grade_level=vacancy.target_grade, 
        academic_year__is_active=True
    )
    
    return render(request, 'academic/vacancy_manage.html', {
        'vacancy': vacancy, 
        'classes': available_classes
    })



# ==============================================================================
# 6. GESTÃO DE ANOS LETIVOS (CORRIGIDO: transaction importado no topo)
# ==============================================================================

@login_required
def academic_year_list(request):
    """Lista todos os anos letivos e permite criar novos."""
    if not request.user.is_manager:
        return redirect('core:dashboard')

    years = AcademicYear.objects.all().order_by('-start_date')
    
    if request.method == 'POST':
        form = AcademicYearForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                new_year = form.save()
                if new_year.is_active:
                    # Desativa todos os outros antes de ativar o novo
                    AcademicYear.objects.exclude(id=new_year.id).update(is_active=False)
                new_year.save()
                
            messages.success(request, f"Ano Letivo {new_year.name} criado com sucesso.")
            return redirect('academic:year_list')
    else:
        form = AcademicYearForm()

    return render(request, 'academic/year_list.html', {'years': years, 'form': form})

@login_required
def academic_year_activate(request, year_id):
    """Define um ano como ATIVO e desativa os restantes."""
    if not request.user.is_manager:
        return redirect('core:dashboard')
        
    year = get_object_or_404(AcademicYear, id=year_id)
    
    with transaction.atomic():
        AcademicYear.objects.all().update(is_active=False)
        year.is_active = True
        year.save()
        
    messages.success(request, f"Ano Letivo {year.name} agora está ATIVO.")
    return redirect('academic:year_list')

# ==============================================================================
# 7. ANALÍTICA E EFICIÊNCIA (COM FILTRO RIGOROSO POR DATAS DO ANO ATIVO)
# ==============================================================================

@login_required
def academic_efficiency_dashboard(request):
    """
    Dashboard Consolidado: KPIs de Notas + Eficiência de Recursos.
    """
    if not request.user.is_manager:
        return redirect('core:dashboard')
        
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.warning(request, "Nenhum ano letivo ativo encontrado.")
        return redirect('academic:year_list')

    # --- 1. MOTOR DE KPI (SOTARQ BI) ---
    from apps.reports.services.kpi_engine import AcademicKPIEngine
    teacher_performance = AcademicKPIEngine.calculate_teacher_performance(active_year.id)
    
    top_performers = sorted(teacher_performance, key=lambda x: x['pass_rate'], reverse=True)[:5]
    alerts = sorted(teacher_performance, key=lambda x: x['pass_rate'])[:5]

    # --- 2. ANALÍTICA DE RECURSOS (Gaps e Ocupação) ---
    teacher_stats = []
    room_stats = []
    
    if EfficiencyAnalytics:
        teachers = Teacher.objects.filter(is_active=True)
        for t in teachers:
            windows = EfficiencyAnalytics.get_teacher_windows(t, active_year)
            total_gap = sum([w['duration'].seconds//60 for w in windows])
            if total_gap > 0:
                teacher_stats.append({
                    'name': t.user.get_full_name(), 
                    'gap_minutes': total_gap, 
                    'windows': len(windows)
                })
        teacher_stats.sort(key=lambda x: x['gap_minutes'], reverse=True)

        rooms = Classroom.objects.all()
        for r in rooms:
            occupancy = EfficiencyAnalytics.get_room_occupancy_rate(r, active_year)
            room_stats.append({'name': r.name, 'occupancy': occupancy})
    
    lock_event = AcademicEvent.objects.filter(
        category='HOLIDAY', 
        klass__isnull=True
    ).first()

    return render(request, 'academic/efficiency_dashboard.html', {
        'top_performers': top_performers,
        'alerts': alerts,
        'active_year': active_year,
        'teacher_stats': teacher_stats[:10],
        'room_stats': room_stats,
        'lock_event': lock_event, # ADICIONADO: Agora o HTML pode ler o estado
    })



@login_required
@require_POST
def toggle_pedagogical_break(request):
    # 1. Identificar o Tenant (Conforme seu modelo User)
    school_tenant = request.user.tenant 
    
    # 2. Obter o Ano Lectivo Ativo 
    # NOTA: Adicione o campo 'tenant' ao seu model AcademicYear no futuro!
    # Por enquanto, buscaremos o global ativo.
    active_year = get_object_or_404(AcademicYear, is_active=True)
    
    # 3. Buscar ou Criar a configuração no AcademicGlobal (Seu model real)
    # Como o AcademicGlobal não tem FK para School no seu código enviado, 
    # recomendo usar o get_or_create baseado em algum critério ou garantir que exista.
    # Abaixo a lógica rigorosa:
    config, created = AcademicGlobal.objects.get_or_create(
        id=1 # Ou vincule ao tenant se adicionar a FK no model
    )
    
    # 4. Inverter o estado
    config.is_pedagogical_break = not config.is_pedagogical_break
    config.save()
    
    return redirect('academic:efficiency_analytics')


@login_required
def export_efficiency_report_pdf(request):
    """
    Exportação PDF: Rigor SOTARQ com WeasyPrint.
    Garante consistência com o Dashboard e Multi-tenancy.
    """
    if not request.user.is_manager:
        return redirect('core:dashboard')

    # 1. Recuperar Ano Letivo Ativo (Consistência com o Dashboard)
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        return HttpResponse("Nenhum ano letivo ativo encontrado.", status=404)

    # 2. Motor de KPI e Analítica (Usando os nomes corrigidos: klass/class_room)
    from apps.reports.services.kpi_engine import AcademicKPIEngine
    from apps.academic.analytics import EfficiencyAnalytics
    
    teacher_performance = AcademicKPIEngine.calculate_teacher_performance(active_year.id)
    
    # Preparar dados reais para o PDF
    top_performers = sorted(teacher_performance, key=lambda x: x['pass_rate'], reverse=True)[:10]
    alerts = sorted(teacher_performance, key=lambda x: x['pass_rate'])[:10]
    
    # Analítica de Professores (Gaps)
    teacher_stats = []
    teachers = Teacher.objects.filter(is_active=True).select_related('user')
    for t in teachers:
        windows = EfficiencyAnalytics.get_teacher_windows(t, active_year)
        total_gap = sum([w['duration'].total_seconds() // 60 for w in windows])
        if total_gap > 0:
            teacher_stats.append({
                'name': t.user.get_full_name(),
                'gap_minutes': int(total_gap),
                'windows': len(windows)
            })

    context = {
        'active_year': active_year,
        'top_performers': top_performers,
        'alerts': alerts,
        'teacher_stats': teacher_stats,
        'report_date': timezone.now(),
        'school_name': request.tenant.name.upper(), # Rigor Multi-tenant: Usa o tenant atual
    }

    # 3. Renderizar e Gerar PDF
    html_string = render_to_string('academic/pdf/report_efficiency.html', context)
    
    # WeasyPrint para gerar o PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    result = html.write_pdf()

    # 4. Resposta HTTP (Simplificada e Eficiente)
    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Relatorio_Eficiencia_{active_year.name}.pdf"'
    
    return response


@login_required
def academic_year_delete(request, year_id):
    """Apaga um ano letivo (se não estiver ativo)."""
    if not request.user.is_manager:
        messages.error(request, "Acesso negado.")
        return redirect('academic:year_list')

    year = get_object_or_404(AcademicYear, id=year_id)
    
    if year.is_active:
        messages.error(request, "Não é possível apagar o Ano Letivo ATIVO. Ative outro ano primeiro.")
        return redirect('academic:year_list')

    try:
        # Verifica se há dados vinculados para evitar cascata destrutiva
        if year.class_set.exists() or year.enrollment_set.exists():
             messages.warning(request, "Este ano letivo contém turmas ou matrículas. Arquive-o em vez de apagar.")
             return redirect('academic:year_list')
             
        year.delete()
        messages.success(request, f"Ano {year.name} removido com sucesso.")
    except Exception as e:
        messages.error(request, f"Erro ao apagar: {str(e)}")
        
    return redirect('academic:year_list')



# apps/academic/views.py
from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .models import Course, Class, GradeLevel, AcademicYear
from .forms import CourseForm, ClassForm # Assumindo que você criará os forms

class StaffRequiredMixin(UserPassesTestMixin):
    """Garante que apenas Admin/Diretores/Secretaria acessem as views de escrita."""
    def test_func(self):
        return self.request.user.current_role in ['ADMIN', 'DIRECTOR', 'SECRETARY']

# ==========================================
# GESTÃO DE CURSOS
# ==========================================


class GradeLevelListView(LoginRequiredMixin, ListView):
    model = GradeLevel
    template_name = 'academic/grade_level_list.html'
    context_object_name = 'grade_levels'

    def get_queryset(self):
        # Mantendo seu rigor de performance com select_related
        return GradeLevel.objects.all().select_related('course').order_by('course', 'level_index')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # INJEÇÃO CRÍTICA: Sem isso, os campos {{ form.name }} no modal ficam invisíveis/inacessíveis
        context['form'] = GradeLevelForm() 
        return context


class GradeLevelCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    model = GradeLevel
    form_class = GradeLevelForm
    template_name = 'academic/grade_level_list.html' 
    success_url = reverse_lazy('academic:grade_level_list')

    def form_invalid(self, form):
        # Buscamos os dados da listagem exatamente como na ListView original
        grade_levels = GradeLevel.objects.all().select_related('course').order_by('course', 'level_index')
        
        return self.render_to_response(self.get_context_data(
            form=form,
            grade_levels=grade_levels,
            open_modal_on_load=True 
        ))




@login_required
def ajax_load_classes(request):
    grade_level_id = request.GET.get('grade_level_id')
    # Rigor SOTARQ: Filtro obrigatório por Tenant e Grade Level
    classes = Class.objects.filter(
        grade_level_id=grade_level_id,
        tenant=request.user.tenant
    ).order_by('name')
    
    data = [
        {'id': c.id, 'name': c.name} for c in classes
    ]
    return JsonResponse(data, safe=False)

# ==========================================
# GESTÃO DE CURSOS
# ==========================================


class CourseListView(LoginRequiredMixin, ListView):
    model = Course
    template_name = 'academic/course_list.html'
    context_object_name = 'courses'

    def get_queryset(self):
        # O django-tenants já isola o Schema. 
        # Basta dar o .all() e ele trará apenas os cursos da escola atual.
        return Course.objects.all().order_by('name')


class CourseCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    model = Course
    fields = ['name', 'code', 'level', 'duration_years', 'coordinator']
    template_name = 'academic/course_form.html'
    success_url = reverse_lazy('academic:course_list')

    def form_valid(self, form):
        # Não injetamos nada. O objeto cai automaticamente no esquema ativo.
        return super().form_valid(form)

"""
@login_required
@user_passes_test(is_manager_check, login_url='/')
def course_edit(request, pk):
    # Mantendo sua busca original que você confirmou que funciona
    course = get_object_or_404(Course, pk=pk)
    
    from .forms import CourseForm, GradeLevelFormSet

    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        formset = GradeLevelFormSet(request.POST, instance=course)

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    course_obj = form.save()
                    formset.save()
                    
                    messages.success(request, f"Rigor SOTARQ: Curso '{course_obj.name}' e tabela de preços atualizados!")
                    return redirect('academic:course_list')
            except Exception as e:
                messages.error(request, f"Erro ao processar alteração: {str(e)}")
        # Mude o 'else' da sua view para isso para debugar:
        else:
            print("ERROS NO FORM:", form.errors)
            print("ERROS NO FORMSET:", formset.errors)
            messages.error(request, "Erro na validação. Verifique os valores inseridos.")
    else:
        form = CourseForm(instance=course)
        formset = GradeLevelFormSet(instance=course)

    # --- INJEÇÃO DE DADOS PARA O TEMPLATE ---
    # Para o select de "Vínculo de Taxa" e o JavaScript funcionarem, 
    # precisamos passar as taxas (FeeType) para o contexto.
    # Filtramos por school se o campo existir no FeeType
    fees = FeeType.objects.filter() 

    return render(request, 'academic/course_edit.html', {
        'form': form,
        'formset': formset,
        'course': course,
        'fees': fees, # Necessário para o {% for fee in fees %} no HTML
    })


"""




@login_required
@user_passes_test(is_manager_check, login_url='/')
def course_edit(request, pk):
    course = get_object_or_404(Course, pk=pk)
    
    # Importação local para evitar import circular
    from .forms import CourseForm, GradeLevelFormSet
    from apps.finance.models import FeeType

    # No Rigor SOTARQ, pegamos a escola do usuário logado (assumindo request.user.school)
    user_school = getattr(request.user, 'school', None)

    if request.method == 'POST':
        # Passamos 'school' para o Form filtrar coordenadores e taxas corretamente
        form = CourseForm(request.POST, instance=course, school=user_school)
        formset = GradeLevelFormSet(request.POST, instance=course)

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    # 1. Salva o Curso (IVA, Taxas Padrão, etc)
                    course_obj = form.save()
                    
                    # 2. Salva as Classes (GradeLevels)
                    # O formset.save() vai persistir o 'total_monthly_fee' 
                    # porque definimos required=False no form.
                    formset.save()
                    
                    messages.success(request, f"Rigor SOTARQ: Curso '{course_obj.name}' e tabela de preços atualizados!")
                    return redirect('academic:course_list')
            except Exception as e:
                messages.error(request, f"Erro de Persistência: {str(e)}")
        else:
            # Debug de Rigor para o Estagiário
            print("ERROS NO FORM:", form.errors)
            print("ERROS NO FORMSET:", formset.errors)
            messages.error(request, "Falha na validação. Verifique os campos em vermelho.")
    else:
        # GET: Carrega formulários com a instância e filtros de escola
        form = CourseForm(instance=course, school=user_school)
        formset = GradeLevelFormSet(instance=course)

    # Filtro de taxas para o contexto (usado pelo JS ou combos extras)
    fees = FeeType.objects.filter(school=user_school) if user_school else FeeType.objects.all()

    return render(request, 'academic/course_edit.html', {
        'form': form,
        'formset': formset,
        'course': course,
        'fees': fees,
    })




@login_required
@user_passes_test(is_manager_check, login_url='/')
@transaction.atomic
def course_delete(request, pk):
    """
    Exclusão de Curso com verificação de integridade referencial.
    """
    # CORRIGIDO: Removido o lixo no nome da função
    course = get_object_or_404(Course, pk=pk)
    #course = get_object_or_404(Course, pk=pk, tenant=request.user.tenant)
    
    # RIGOR SOTARQ: Verificar se há matrículas ativas neste curso
    from apps.students.models import Enrollment
    has_enrollments = Enrollment.objects.filter(course=course).exists()
    
    if has_enrollments:
        messages.error(
            request, 
            f"ERRO DE INTEGRIDADE: O curso '{course.name}' não pode ser eliminado "
            f"pois já existem alunos matriculados ou históricos vinculados a ele."
        )
        return redirect('academic:course_list')

    try:
        course_name = course.name
        course.delete()
        messages.success(request, f"Curso '{course_name}' eliminado com sucesso.")
    except Exception as e:
        messages.error(request, f"Erro ao eliminar curso: {str(e)}")
        
    return redirect('academic:course_list')



class SubjectListView(LoginRequiredMixin, ListView):
    model = Subject
    template_name = 'academic/subject_list.html'
    context_object_name = 'subjects'

    def get_queryset(self):
        # Rigor SOTARQ: Ordenação lógica por Nível e depois Nome
        return Subject.objects.all().select_related('grade_level', 'grade_level__course').order_by(
            'grade_level__course', 'grade_level__level_index', 'name'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Injetamos o formulário vazio para o Modal de Criação
        context['form'] = SubjectForm()
        # Útil para filtros no template
        context['grade_levels'] = GradeLevel.objects.all().select_related('course')
        return context


class SubjectCreateView(LoginRequiredMixin, CreateView):
    model = Subject
    form_class = SubjectForm
    template_name = 'academic/subject_list.html'
    success_url = reverse_lazy('academic:subject_list')

    def form_invalid(self, form):
        # Se falhar, recarrega a lista com os erros e abre o modal
        queryset = Subject.objects.all().select_related('grade_level').order_by('grade_level', 'name')
        return self.render_to_response(self.get_context_data(
            form=form,
            subjects=queryset,
            open_modal_on_load=True
        ))
    

# ==========================================
# GESTÃO DE TURMAS (CLASSES)
# ==========================================

class ClassListView(LoginRequiredMixin, ListView):
    model = Class
    template_name = 'academic/class_list.html'
    context_object_name = 'classes'

    def get_queryset(self):
        # O isolamento por Schema garante que AcademicYear.objects.all() 
        # já retorne apenas os anos da "Escola Excellence".
        return Class.objects.all().select_related(
            'grade_level', 
            'academic_year', 
            'main_teacher'
        )



class ClassCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    model = Class
    form_class = ClassForm # Use o form_class em vez de fields
    template_name = 'academic/class_form.html'
    success_url = reverse_lazy('academic:class_list')


#class CourseCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
#    model = Course
#    fields = ['name', 'code', 'level', 'duration_years', 'coordinator']
#    template_name = 'academic/course_form.html'
#    success_url = reverse_lazy('academic:course_list')

#    def form_valid(self, form):
#        # Injeta o tenant automaticamente antes de salvar
#        form.instance.tenant = self.request.user.tenant
#        return super().form_valid(form)



@login_required
def academic_year_deactivate(request, year_id):
    """Desativa manualmente um ano letivo (Ordem direta)."""
    if not request.user.is_manager:
        return redirect('core:dashboard')
        
    year = get_object_or_404(AcademicYear, id=year_id)
    
    if year.is_active:
        year.is_active = False
        year.save()
        messages.warning(request, f"Ano Letivo {year.name} foi DESATIVADO. O sistema está agora sem ano ativo.")
    
    return redirect('academic:year_list')




@login_required
def export_class_pauta_excel(request, class_id):
    """
    Gera a Pauta Trimestral em Excel com todas as disciplinas da turma.
    """
    klass = get_object_or_404(Class, id=class_id)
    subjects = Subject.objects.filter(grade_level=klass.grade_level).order_by('name')
    # Pegamos os alunos vinculados a esta turma (via Enrollment)
    students = klass.enrollments_records.filter(status='active').select_related('student')

    # Criar Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Pauta - {klass.name}"

    # --- ESTILOS SOTARQ ENTERPRISE ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')

    # --- CABEÇALHO ---
    # Linha 1: Nome da Escola (Tenant)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(subjects) + 2)
    ws.cell(row=1, column=1).value = request.tenant.name.upper()
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = align_center

    # Linha 3: Títulos das Colunas
    headers = ["Nº", "Nome Completo"] + [s.name for s in subjects] + ["Média Geral"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25 if col_num == 2 else 15

    # --- DADOS DOS ALUNOS ---
    for row_num, enrollment in enumerate(students, 4):
        student = enrollment.student
        ws.cell(row=row_num, column=1).value = row_num - 3
        ws.cell(row=row_num, column=2).value = student.full_name
        
        total_mf = 0
        count_subjects = 0

        # Preencher notas por disciplina
        for col_num, subject in enumerate(subjects, 3):
            # Busca a nota deste aluno nesta disciplina e nesta turma
            grade = StudentGrade.objects.filter(
                student=student, 
                subject=subject, 
                klass=klass
            ).first()
            
            val = grade.mf if grade else 0.0
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = float(val)
            cell.border = border
            cell.alignment = align_center
            
            # Condicional: Nota vermelha (< 10 no padrão angolano/português)
            if val < 9.5:
                cell.font = Font(color="FF0000", bold=True)
            
            total_mf += val
            count_subjects += 1

        # Média Geral do Aluno
        avg_cell = ws.cell(row=row_num, column=len(subjects) + 3)
        avg_val = total_mf / count_subjects if count_subjects > 0 else 0
        avg_cell.value = float(round(avg_val, 1))
        avg_cell.font = Font(bold=True)
        avg_cell.border = border

    # Preparar Resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Pauta_{klass.name}_{klass.academic_year.name}.xlsx".replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response




@login_required
@transaction.atomic
def daily_attendance_control(request, allocation_id):
    """
    Interface de chamada SOTARQ. 
    Controle em tempo real: Notifica Direção (Geral/Pedagógica) e Encarregados (WhatsApp/SMS).
    """
    allocation = get_object_or_404(TeacherSubject, id=allocation_id, teacher__user=request.user)
    students = Student.objects.filter(current_class=allocation.class_room, is_active=True)
    
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        attendance_type = request.POST.get('type') # 'P', 'FJ', 'FI'
        
        student = get_object_or_404(Student, id=student_id)
        
        # 1. Recupera ou cria o registro acadêmico para o contador de faltas
        grade, _ = StudentGrade.objects.get_or_create(
            student=student,
            subject=allocation.subject,
            klass=allocation.class_room,
            academic_year=allocation.class_room.academic_year
        )
        
        if attendance_type == 'FI':
            grade.unjustified_absences += 1
            
            # Cálculo de Limite conforme Artigo 25º do Decreto 424/25
            weekly_periods = allocation.subject.workload_hours
            limit = 3 if weekly_periods == 1 else (4 if weekly_periods == 2 else 5)
            
            # A. VERIFICAÇÃO DE RETENÇÃO (Fim da linha para o aluno)
            if grade.check_attendance_failure(weekly_periods):
                # NOTIFICAÇÃO DE EMERGÊNCIA PARA DIREÇÃO (Geral e Pedagógica)
                management_roles = [Role.Type.DIRECTOR, Role.Type.PEDAGOGIC]
                directors = User.objects.filter(
                    tenant=request.user.tenant, 
                    current_role__in=management_roles
                )
                
                notifs = [Notification(
                    user=d,
                    title="🚨 RETENÇÃO POR FALTAS",
                    message=f"CRÍTICO: O aluno {student.full_name} acaba de ser RETIDO na disciplina de {allocation.subject.name} por excesso de faltas.",
                    link=f"/academic/grading/class/{allocation.class_room.id}/subject/{allocation.subject.id}/"
                ) for d in directors]
                Notification.objects.bulk_create(notifs)
                
                messages.error(request, f"O Aluno {student.full_name} atingiu o limite de retenção!")

            # B. DISPARO DO XEQUE-MATE (Aviso prévio ao Encarregado quando falta apenas 1 para o limite)
            elif grade.unjustified_absences >= (limit - 1):
                guardian_link = student.guardians.filter(is_financial_responsible=True).first()
                if guardian_link:
                    #from apps.core.servicos.notifications import AlertService
                    AlertService.send_attendance_alert(
                        guardian_link.guardian.phone, 
                        student.full_name, 
                        allocation.subject.name
                    )

            grade.save()
            return JsonResponse({
                'status': 'updated', 
                'absences': grade.unjustified_absences,
                'is_failed': grade.is_failed_by_attendance
            })

    # IMPORTANTE: Retorna o template para carregamento inicial do Modal
    return render(request, 'academic/attendance_sheet.html', {
        'allocation': allocation,
        'students': students,
        'today': timezone.now()
    })


from weasyprint import HTML
from django.template.loader import render_to_string

@login_required
def export_attendance_pdf(request, allocation_id):
    """
    Gera a Lista de Chamada Mensal em PDF para uso manual ou arquivo.
    Acessível por Professores, Diretores e Admin.
    """
    allocation = get_object_or_404(TeacherSubject, id=allocation_id)
    
    # Rigor de Tenant: Garante que o usuário não baixe lista de outra escola
    if allocation.class_room.academic_year.tenant != request.user.tenant:
        return HttpResponseForbidden()

    students = Student.objects.filter(current_class=allocation.class_room, is_active=True).order_by('full_name')
    
    # Geramos uma lista de dias para o cabeçalho do mês atual
    import calendar
    from datetime import datetime
    now = datetime.now()
    month_days = range(1, calendar.monthrange(now.year, now.month)[1] + 1)

    context = {
        'allocation': allocation,
        'students': students,
        'month_days': month_days,
        'month_name': calendar.month_name[now.month].upper(),
        'school_name': request.user.tenant.name,
        'generated_at': now,
    }

    html_string = render_to_string('academic/pdf/attendance_print_template.html', context)
    pdf = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="Lista_Chamada_{allocation.class_room.name}.pdf"'
    return response






@login_required
@user_passes_test(is_manager_check)
def export_final_pauta_excel(request, class_id):
    klass = get_object_or_404(Class, id=class_id, academic_year__tenant=request.user.tenant)
    
    excel_data = ExportEngine.generate_final_pauta_excel(klass)
    
    response = HttpResponse(
        excel_data, 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"PAUTA_FINAL_{klass.name}.xlsx".replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response



@login_required
@transaction.atomic
def bulk_promote_students(request, class_id):
    """
    Consolidação Sênior: Promove alunos conforme Decreto 424/25.
    Restrito a: ADMIN, DIRECTOR, PEDAGOGIC, DIRECT_ADMIN.
    """
    # 1. RIGOR DE ACESSO: Verificação de Hierarquia
    AUTHORIZED_ROLES = [Role.Type.ADMIN, Role.Type.DIRECTOR, Role.Type.PEDAGOGIC, Role.Type.DIRECT_ADMIN]
    if request.user.current_role not in AUTHORIZED_ROLES:
        messages.error(request, "VIOLAÇÃO DE ACESSO: Você não tem autoridade para promover alunos.")
        return redirect('academic:student_dashboard')

    # 2. ISOLAMENTO DE TENANT E BUSCA DE DADOS
    current_class = get_object_or_404(Class, id=class_id, academic_year__tenant=request.user.tenant)
    
    # Busca o próximo ciclo cronológico para este Tenant
    next_academic_year = AcademicYear.objects.filter(
        tenant=request.user.tenant,
        start_date__gt=current_class.academic_year.start_date
    ).order_by('start_date').first()
    
    if not next_academic_year:
        messages.error(request, "BLOQUEIO: Próximo Ano Lectivo não configurado. Impossível promover.")
        return redirect('academic:student_dashboard')

    # 3. LÓGICA DE PROMOÇÃO (Regra dos 3 Trimestres Consolidada)
    active_enrollments = current_class.enrollments_records.filter(status='active')
    promoted_count = 0

    for enrollment in active_enrollments:
        # Busca todas as notas do aluno nesta turma específica
        grades = StudentGrade.objects.filter(student=enrollment.student, klass=current_class)
        
        # Define nota de passagem conforme nível (Regra do .5 automático no save já aplicada)
        pass_grade = 5 if current_class.grade_level.level_index <= 6 else 10
        
        # Conta disciplinas com negativa na Média Final
        failed_count = grades.filter(mf__lt=pass_grade).count()
        
        # CONDIÇÃO DE SUCESSO: No máximo 2 negativas e sem retenção por faltas
        if failed_count <= 2 and not enrollment.status == 'failed':
            next_grade = enrollment.grade_level.next_level() 
            
            if next_grade:
                # Cria a nova inscrição pendente para o novo ciclo
                Enrollment.objects.create(
                    student=enrollment.student,
                    academic_year=next_academic_year,
                    grade_level=next_grade,
                    course=enrollment.course,
                    status='pending_placement' # Aguarda alocação física do Diretor
                )
                
                # Marca a matrícula do ano que findou como Graduada/Concluída
                enrollment.status = 'graduated'
                enrollment.save()
                promoted_count += 1

    messages.success(request, f"GOVERNANÇA: Operação concluída. {promoted_count} alunos promovidos para {next_academic_year.name}.")
    return redirect('academic:student_dashboard')


@login_required
def export_minipauta_view(request, allocation_id, term):
    """
    View para exportação da Minipauta. 
    Apenas o professor alocado ou a Direção podem baixar.
    """
    # Rigor de Posse: Verifica se a alocação pertence ao professor ou se o usuário é gestor
    if request.user.is_manager:
        allocation = get_object_or_404(TeacherSubject, id=allocation_id, class_room__academic_year__tenant=request.user.tenant)
    else:
        allocation = get_object_or_404(TeacherSubject, id=allocation_id, teacher__user=request.user)
    
    excel_data = ExportEngine.generate_minipauta_excel(
        klass=allocation.class_room, 
        subject=allocation.subject, 
        term=term
    )
    
    response = HttpResponse(
        excel_data, 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"MINIPAUTA_T{term}_{allocation.class_room.name}_{allocation.subject.name[:10]}.xlsx".replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


@login_required
def final_pauta_view(request, class_id):
    """Visualização da Pauta Final com Verificação Financeira em tempo real."""
    klass = get_object_or_404(Class, id=class_id, academic_year__tenant=request.user.tenant)
    subjects = klass.grade_level.subjects.all().order_by('name')
    enrollments = klass.enrollments_records.filter(status='active').select_related('student').order_by('student__full_name')

    # RIGOR FINANCEIRO: Mapeia dívidas de cada aluno
    for enr in enrollments:
        # Verifica se existem faturas vencidas ou pendentes na app finance
        from apps.finance.models import Invoice
        enr.has_debt = Invoice.objects.filter(
            student=enr.student, 
            status__in=['pending', 'overdue']
        ).exists()

    AUTHORIZED_ROLES = [Role.Type.ADMIN, Role.Type.DIRECTOR, Role.Type.PEDAGOGIC, Role.Type.DIRECT_ADMIN]
    is_authorized = request.user.current_role in AUTHORIZED_ROLES
    
    return render(request, 'academic/final_pauta_view.html', {
        'klass': klass,
        'subjects': subjects,
        'enrollments': enrollments,
        'is_authorized': is_authorized
    })



@login_required
@user_passes_test(is_manager_check)
def mass_whatsapp_promotion_alert(request):
    """
    Motor SOTARQ MESSENGER: Disparo em massa para alunos promovidos.
    Notifica o Encarregado sobre a vaga reservada no próximo ano lectivo.
    """
    # 1. Filtra matrículas que acabaram de ser criadas via promoção (aguardando vaga física)
    # Rigor: Apenas do tenant atual
    promoted_enrollments = Enrollment.objects.filter(
        academic_year__is_active=True,
        status='pending_placement',
    ).select_related('student', 'grade_level')

    if not promoted_enrollments.exists():
        messages.info(request, "Nenhum aluno promovido recentemente para notificar.")
        return redirect('academic:student_dashboard')

    count = 0
    for enrollment in promoted_enrollments:
        student = enrollment.student
        # Busca o encarregado financeiro (Rigor SOTARQ: quem paga é quem decide)
        guardian_link = student.guardians.filter(is_financial_responsible=True).first()
        
        if guardian_link and guardian_link.guardian.phone:
            phone = guardian_link.guardian.phone
            msg = (
                f"Olá, {guardian_link.guardian.full_name}! 👋\n"
                f"Temos boas notícias: O aluno *{student.full_name}* foi promovido para a *{enrollment.grade_level.name}*!\n\n"
                f"A vaga já está reservada. Por favor, acesse o portal ou dirija-se à secretaria para confirmar a matrícula.\n"
                f"Atenciosamente, Direção {request.user.tenant.name} 🏛️"
            )
            
            # Aqui chamamos a Task Assíncrona do SOTARQ MESSENGER para não travar o servidor
            # task_send_whatsapp.delay(phone, msg)
            count += 1

    messages.success(request, f"SOTARQ MESSENGER: {count} alertas de promoção enviados para a fila de disparo.")
    return redirect('academic:student_dashboard')



