# apps/core/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden
from django.db import connections
from django.db import transaction
from django.contrib.auth.views import LoginView
from django.shortcuts import resolve_url
from apps.cafeteria.services import CafeteriaInventoryService
from apps.core.views_admin import is_manager_check
from apps.academic.models import Course
from apps.fiscal.models import DocumentoCanceladoAudit, DocumentoFiscal
from apps.inventory.services import AssetManager
from .models import Notification, HelpArticle, SchoolConfiguration, SupportTicket, UserRole
import json
from django.utils import timezone
from django.db.models import Sum, Count, Q
from decimal import Decimal
from apps.students.models import Student
from apps.finance.models import Invoice, Payment
from django.template.loader import render_to_string
from dateutil.relativedelta import relativedelta
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from .models import User, Role 
import io
from django.db import transaction, IntegrityError
from apps.teachers.models import Teacher
from .forms import SchoolSettingsForm, UserImportForm, UserManagementForm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone




# Views que controla as notificações
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Notification

@login_required
def check_new_notifications(request):
    """
    Endpoint para o Polling AJAX do base.html.
    Verifica se há novas notificações desde o último check.
    """
    since = request.GET.get('since')
    
    # Busca novas notificações não lidas para o usuário atual
    # Filtro rigoroso por usuário para evitar vazamento de dados entre schemas
    new_notifs = Notification.objects.filter(
        user=request.user,
        is_read=False,
        created_at__gt=since
    ).order_by('-created_at')

    total_unread = Notification.objects.filter(user=request.user, is_read=False).count()

    latest = None
    if new_notifs.exists():
        top = new_notifs.first()
        latest = {
            'title': top.title,
            'message': top.message,
            'link': top.link or '#'
        }

    return JsonResponse({
        'count': new_notifs.count(),
        'total_unread': total_unread,
        'latest': latest
    })


########################################
# Área do site institucional
########################################

class CustomLoginView(LoginView):
    template_name = 'core/login.html' # Seu template de login

    def get_success_url(self):
        """
        Lógica de Redirecionamento Inteligente SOTARQ.
        """
        user = self.request.user
        role = user.current_role
        
        # 1. Alunos e Encarregados -> Portal do Aluno
        if role in ['STUDENT', 'GUARDIAN']:
            return resolve_url('portal:dashboard')
        
        # 2. Staff (Admin, Diretor, Prof, Sec, RH, Contab) -> ERP Dashboard
        # Eles acessam o portal via botão no menu, se permitido.
        return resolve_url('core:dashboard')

# Views Institucionais (Site Público)
def public_about(request):
    return render(request, 'public/about.html')

def public_courses(request):
    courses = Course.objects.filter(deleted_at__isnull=True).order_by('name')
    return render(request, 'public/courses.html', {'courses': courses})

def public_contact(request):
    return render(request, 'public/contact.html')




@login_required
def school_configuration_update(request):
    """
    Painel onde o Diretor edita a aparência e configurações da escola.
    """
    # Apenas Diretor/Admin
    if request.user.current_role not in [Role.Type.ADMIN, Role.Type.DIRECTOR]:
        return redirect('core:dashboard')

    # Garante que existe uma configuração
    config, created = SchoolConfiguration.objects.get_or_create(id=1) # Assume single-tenant logic per schema

    if request.method == 'POST':
        form = SchoolSettingsForm(request.POST, request.FILES, instance=config)
        if form.is_valid():
            form.save()
            # Feedback visual de sucesso
            from django.contrib import messages
            messages.success(request, "Configurações da escola atualizadas com sucesso!")
            return redirect('core:school_settings')
    else:
        form = SchoolSettingsForm(instance=config)

    return render(request, 'core/settings/school_settings.html', {'form': form})







@login_required
def dashboard(request):
    """
    PAINEL EXECUTIVO K12: Centraliza métricas da escola única do Tenant.
    A filtragem por schema já garante o isolamento dos dados.
    """
    today = timezone.now().date()
    
    # --- GRÁFICO DE RECEITA (6 MESES) ---
    data_receita = []
    labels_meses = []
    meses_pt = {1:'Jan', 2:'Fev', 3:'Mar', 4:'Abr', 5:'Mai', 6:'Jun',
                7:'Jul', 8:'Ago', 9:'Set', 10:'Out', 11:'Nov', 12:'Dez'}

    for i in range(5, -1, -1):
        date_ref = today - relativedelta(months=i)
        mes, ano = date_ref.month, date_ref.year
        
        # Filtro global no schema: Removido 'school_unit=unit'
        total_mes = Payment.objects.filter(
            confirmed_at__month=mes,
            confirmed_at__year=ano,
            validation_status='validated'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        data_rece_float = float(total_mes)
        data_receita.append(data_rece_float)
        labels_meses.append(f"{meses_pt[mes]}/{str(ano)[2:]}")

    # --- MÉTRICAS E KPIs ---
    faturas_atrasadas = Invoice.objects.filter(
        status__in=['pending', 'overdue'],
        due_date__lt=today
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')

    valuation = AssetManager.get_patrimony_valuation() 
    
    # Busca dados da App Cafeteria
    total_inventory_cost = CafeteriaInventoryService.get_total_cost()
    critical_items_count = CafeteriaInventoryService.check_stock_alerts().count()

    context = {
        'page_title': 'Análise Executiva SOTARQ',
        'total_alunos': Student.objects.filter(is_active=True).count(),
        'receita_mes': data_receita[-1],
        'faturas_atrasadas': faturas_atrasadas,
        'tickets_abertos': SupportTicket.objects.filter(status='open').count(),
        'notificacoes_count': Notification.objects.filter(user=request.user, is_read=False).count(),
        'chart_receita_data': json.dumps(data_receita),
        'chart_receita_labels': json.dumps(labels_meses),
        'chart_pie_data': json.dumps([
            int(Invoice.objects.filter(status='paid').count()),
            int(Invoice.objects.filter(status='overdue').count()),
            int(Invoice.objects.filter(status='pending').count()),
        ]),
        'valuation': valuation, # Dicionário com purchase_value e total_depreciation
        'total_inventory_cost': total_inventory_cost,
        'critical_items_count': critical_items_count,
    }
    return render(request, 'core/dashboard.html', context)



def logout_view(request):
    """
    SOTARQ Custom Logout: Aceita GET e POST.
    Resolve o erro 405 do Django 5+.
    """
    auth_logout(request)
    messages.info(request, "Sessão terminada com sucesso. Até à próxima!")
    return redirect('core:login')


def get_notifications(request):
    """HTMX parcial para o sininho."""
    if not request.user.is_authenticated:
        return HttpResponse("")
    notifications = Notification.objects.filter(user=request.user, is_read=False)[:10]
    return render(request, 'core/partials/notification_list.html', {'notifications': notifications})

@staff_member_required
def help_center(request):
    """Base de conhecimento para funcionários."""
    articles = HelpArticle.objects.filter(deleted_at__isnull=True).order_by('category')
    return render(request, 'core/help_center.html', {'articles': articles})

def health_check(request):
    """Monitor de integridade técnica."""
    try:
        cursor = connections['default'].cursor()
        cursor.execute("SELECT 1")
        return HttpResponse("OK", status=200)
    except Exception:
        return HttpResponse("Database Unavailable", status=503)

@login_required
def export_dashboard_pdf(request):
    """Gera PDF executivo focado no Tenant atual."""
    from weasyprint import HTML
    today = timezone.now()
    
    # KPIs globais do schema
    context = {
        'school_name': request.tenant.name,
        'today': today,
        'total_alunos': Student.objects.filter(is_active=True).count(),
        'receita_mes': Payment.objects.filter(
            confirmed_at__month=today.month,
            validation_status='validated'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
        'faturas_atrasadas': Invoice.objects.filter(
            status__in=['pending', 'overdue'],
            due_date__lt=today.date()
        ).aggregate(total=Sum('total'))['total'] or Decimal('0.00'),
    }

    html_string = render_to_string('core/reports/dashboard_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    result = html.write_pdf()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Relatorio_Executivo_{request.tenant.schema_name}.pdf"'
    response.write(result)
    
    return response




# --- Função Auxiliar de Verificação de Permissão ---
def is_manager_check(user):
    """
    Verifica se o utilizador é Superuser, Admin ou Diretor.
    Utilizado pelo decorador @user_passes_test.
    """
    if not user.is_authenticated:
        return False
        
    # Verifica Superuser
    if user.is_superuser:
        return True
        
    # Verifica Roles de Gestão (Baseado no seu modelo User e Role)
    MANAGEMENT_ROLES = ['ADMIN', 'DIRECTOR'] # Códigos definidos em Role.Type
    return user.current_role in MANAGEMENT_ROLES

# --- View de Gestão de Usuários ---
@login_required
@user_passes_test(is_manager_check, login_url='/', redirect_field_name=None)
def user_management_list(request):
    """
    Lista os funcionários do Tenant atual.
    Filtra para mostrar apenas staff relevante (Admin, Diretor, Professor, Secretaria).
    """
    # Roles que devem aparecer na lista (exclui alunos/encarregados para limpeza visual)
    staff_roles = ['ADMIN', 'DIRECTOR', 'TEACHER', 'SECRETARY']
    
    # Captura filtros da URL (Query Params)
    search_query = request.GET.get('q', '')
    role_filter = request.GET.get('role', '')

    # Query Base: Apenas utilizadores da escola (tenant) atual e com roles de staff
    users = User.objects.filter(
        tenant=request.user.tenant,
        current_role__in=staff_roles
    ).select_related('tenant').order_by('-date_joined')

    # Aplicação de Filtros de Busca
    if search_query:
        from django.db.models import Q
        users = users.filter(
            Q(username__icontains=search_query) | 
            Q(first_name__icontains=search_query) | 
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # Aplicação de Filtro de Role Específico
    if role_filter:
        users = users.filter(current_role=role_filter)

    return render(request, 'core/user_management_list.html', {
        'users': users,
        'page_title': 'Gestão de Usuários'
    })



@login_required
@user_passes_test(is_manager_check, login_url='/', redirect_field_name=None)
def user_add(request):
    if request.method == 'POST':
        form = UserManagementForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. Cria o Usuário mas não salva M2M ainda
                    user = form.save(commit=False)
                    user.tenant = request.user.tenant # Vincula à escola atual
                    
                    # Define senha
                    password = form.cleaned_data.get('password')
                    if password:
                        user.set_password(password)
                    else:
                        user.set_password('Sotarq.123') # Senha padrão se vazia
                    
                    user.save()

                    # 2. Sincroniza a Tabela de Roles (Role M2M)
                    role_code = form.cleaned_data.get('current_role')
                    
                    # Busca ou Cria o Role no Banco (safety check)
                    role_obj, _ = Role.objects.get_or_create(
                        code=role_code, 
                        defaults={'name': role_code}
                    )
                    
                    # Cria o vínculo explícito
                    UserRole.objects.create(user=user, role=role_obj)

                    messages.success(request, f"Funcionário {user.username} criado com sucesso.")
                    return redirect('core:user_management_list')
            except Exception as e:
                messages.error(request, f"Erro ao criar usuário: {str(e)}")
    else:
        form = UserManagementForm()

    return render(request, 'core/user_form.html', {
        'form': form,
        'title': 'Novo Funcionário'
    })

@login_required
@user_passes_test(is_manager_check, login_url='/', redirect_field_name=None)
def user_edit(request, user_id):
    # Garante que só edita usuários do mesmo tenant
    user_obj = get_object_or_404(User, id=user_id, tenant=request.user.tenant)
    
    if request.method == 'POST':
        form = UserManagementForm(request.POST, instance=user_obj)
        if form.is_valid():
            with transaction.atomic():
                user = form.save(commit=False)
                
                # Atualiza senha apenas se fornecida
                password = form.cleaned_data.get('password')
                if password:
                    user.set_password(password)
                
                user.save()
                
                # Atualiza Role (Se mudou)
                role_code = form.cleaned_data.get('current_role')
                role_obj, _ = Role.objects.get_or_create(code=role_code, defaults={'name': role_code})
                
                # Limpa roles anteriores e define o novo (Simples para gestão escolar)
                UserRole.objects.filter(user=user).delete()
                UserRole.objects.create(user=user, role=role_obj)

                messages.success(request, "Dados atualizados com sucesso.")
                return redirect('core:user_management_list')
    else:
        form = UserManagementForm(instance=user_obj)

    return render(request, 'core/user_form.html', {
        'form': form,
        'title': f'Editar {user_obj.username}'
    })




@login_required
@user_passes_test(is_manager_check, login_url='/')
def user_export_excel(request):
    """
    Exporta lista de usuários em formato Excel Nativo (.xlsx) com formatação Enterprise.
    """
    # 1. Configuração do Workbook e Worksheet
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Staff_SOTARQ.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Funcionários"

    # 2. Reaplica os filtros (Mesma lógica segura do CSV)
    staff_roles = ['ADMIN', 'DIRECTOR', 'TEACHER', 'SECRETARY']
    search_query = request.GET.get('q', '')
    role_filter = request.GET.get('role', '')

    users = User.objects.filter(
        tenant=request.user.tenant,
        current_role__in=staff_roles
    ).select_related('teacher_profile').order_by('-date_joined')

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) | 
            Q(first_name__icontains=search_query) | 
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    if role_filter:
        users = users.filter(current_role=role_filter)

    # 3. Definição do Cabeçalho e Estilos
    headers = [
        'ID', 'Username', 'Nome Completo', 'Email Institucional', 
        'Função (Role)', 'Nº Funcionário', 'Grau Académico', 
        'Status', 'Data Cadastro'
    ]
    
    # Estilo do Cabeçalho: Fundo Azul (Brand Color), Texto Branco, Negrito
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5") # Cor Primary do seu CSS
    alignment_center = Alignment(horizontal="center", vertical="center")
    
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # 4. Preenchimento dos Dados
    for user in users:
        # Recupera dados do professor de forma segura
        emp_num = "N/A"
        degree = "N/A"
        
        # Verifica se é professor e tem perfil
        if user.current_role == 'TEACHER' and hasattr(user, 'teacher_profile'):
            emp_num = user.teacher_profile.employee_number
            degree = user.teacher_profile.academic_degree
        
        status = "Ativo" if user.is_active else "Inativo"
        date_joined = user.date_joined.strftime('%d/%m/%Y') if user.date_joined else ""

        row = [
            user.id,
            user.username,
            user.get_full_name(),
            user.email,
            user.get_current_role_display(),
            emp_num,
            degree,
            status,
            date_joined
        ]
        ws.append(row)

    # 5. Ajuste Automático da Largura das Colunas
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4 # +4 para respiro visual

    # Salva o arquivo na resposta HTTP
    wb.save(response)
    return response




@login_required
@user_passes_test(is_manager_check, login_url='/')
def user_download_import_template(request):
    """
    Gera e serve o modelo de Excel (.xlsx) para importação de funcionários.
    Inclui cabeçalhos obrigatórios e uma linha de exemplo.
    """
    # 1. Configuração do Workbook
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Modelo_Importacao_Funcionarios.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo de Importação"

    # 2. Definição das Colunas (Sincronizado com a View de Importação)
    headers = [
        'username', 
        'first_name', 
        'last_name', 
        'email', 
        'role', 
        'employee_number', 
        'academic_degree'
    ]

    # 3. Estilização Enterprise (Cabeçalho Azul e Negrito)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5") # Brand Primary Color
    center_align = Alignment(horizontal='left', vertical='center')

    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 4. Adicionar Dados de Exemplo (Para guiar o utilizador)
    example_rows = [
        # Exemplo Professor
        ['joao.silva', 'João', 'Silva', 'joao.silva@escola.ao', 'TEACHER', 'DOC-2026-001', 'Licenciatura em Matemática'],
        # Exemplo Secretária
        ['ana.costa', 'Ana', 'Costa', 'ana.costa@escola.ao', 'SECRETARY', '', '']
    ]

    for row in example_rows:
        ws.append(row)

    # 5. Adicionar Notas Explicativas (Comentários Visuais na linha 4)
    # Mesclar células para instruções
    ws.merge_cells('A4:G4')
    instruction_cell = ws['A4']
    instruction_cell.value = "NOTA: O campo 'role' deve ser APENAS 'TEACHER' ou 'SECRETARY'. O 'employee_number' é obrigatório apenas para professores."
    instruction_cell.font = Font(italic=True, color="FF0000", size=10)
    
    # 6. Ajuste Automático de Largura
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 5

    wb.save(response)
    return response


@login_required
@user_passes_test(is_manager_check, login_url='/')
def user_import_bulk(request):
    """
    Importação em massa de usuários via Excel (.xlsx).
    Utiliza transações atômicas para garantir a integridade dos dados relacionais.
    """
    if request.method == 'POST':
        form = UserImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            
            try:
                # Carrega o workbook. data_only=True garante que pegamos o valor e não fórmulas.
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active # Pega a primeira aba
                
                # Conversão das linhas para lista para facilitar manipulação
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                    messages.error(request, "O arquivo está vazio.")
                    return redirect('core:user_import')

                # Cabeçalho (Primeira linha) - Normaliza para lower case para mapeamento seguro
                header_row = [str(cell).strip().lower() for cell in rows[0] if cell is not None]
                
                # Mapa de colunas obrigatórias
                required_cols = ['username', 'email', 'role']
                
                # Validação de Cabeçalho
                missing_cols = [col for col in required_cols if col not in header_row]
                if missing_cols:
                    messages.error(request, f"Colunas obrigatórias ausentes: {', '.join(missing_cols)}")
                    return redirect('core:user_import')

                # Mapeia o índice de cada coluna (ex: 'email' está na coluna 3)
                col_map = {name: index for index, name in enumerate(header_row)}

                # Limite de segurança enterprise (excluindo cabeçalho)
                # Limite de importação simultânea
                data_rows = rows[1:]
                if len(data_rows) > 500:
                    messages.error(request, "O arquivo excede o limite de 500 registros por vez.")
                    return redirect('core:user_import')

                success_count = 0
                errors = []

                for row_idx, row in enumerate(data_rows, start=2): # start=2 pois linha 1 é cabeçalho
                    # Função auxiliar segura para pegar valor da célula pelo nome da coluna
                    def get_val(col_name):
                        if col_name in col_map and col_map[col_name] < len(row):
                            val = row[col_map[col_name]]
                            return str(val).strip() if val is not None else ''
                        return ''

                    try:
                        # Extração de dados usando o mapa
                        username = get_val('username')
                        email = get_val('email')
                        role_code = get_val('role').upper()
                        first_name = get_val('first_name')
                        last_name = get_val('last_name')
                        
                        # Pula linhas vazias (comum em Excel)
                        if not username and not email:
                            continue

                        # Validação de Dados Básicos
                        if not username or not email or role_code not in ['TEACHER', 'SECRETARY']:
                            raise ValueError(f"Dados obrigatórios ausentes ou Role inválido (Use TEACHER ou SECRETARY).")

                        with transaction.atomic():
                            # 1. Cria ou Verifica Usuário
                            if User.objects.filter(username=username).exists():
                                raise IntegrityError(f"Username '{username}' já existe.")
                            
                            user = User.objects.create_user(
                                username=username,
                                email=email,
                                password='Sotarq.ChangeMe', # Senha padrão
                                first_name=first_name,
                                last_name=last_name,
                                tenant=request.user.tenant,
                                current_role=role_code,
                                is_active=True
                            )

                            # 2. Atribui Role e Cria Vínculo M2M
                            role_obj, _ = Role.objects.get_or_create(code=role_code, defaults={'name': role_code})
                            UserRole.objects.create(user=user, role=role_obj)

                            # 3. Lógica específica para Professores
                            if role_code == 'TEACHER':
                                emp_number = get_val('employee_number')
                                degree = get_val('academic_degree')
                                
                                if not emp_number:
                                    raise ValueError("Nº Funcionário é obrigatório para Professores.")
                                    
                                Teacher.objects.create(
                                    user=user,
                                    employee_number=emp_number,
                                    academic_degree=degree or 'N/A'
                                )

                            success_count += 1
                            
                    except Exception as e:
                        errors.append(f"Linha {row_idx} ({get_val('username') or 'Desconhecido'}): {str(e)}")

                # Feedback ao usuário
                if success_count > 0:
                    messages.success(request, f"{success_count} funcionários importados com sucesso.")
                
                if errors:
                    for err in errors[:5]:
                        messages.error(request, err)
                    if len(errors) > 5:
                        messages.warning(request, f"E mais {len(errors) - 5} erros não listados.")
                
                if success_count > 0 and not errors:
                     return redirect('core:user_management_list')

            except Exception as e:
                # Erro genérico ao abrir o arquivo (formato inválido, corrompido, etc)
                messages.error(request, f"Erro crítico ao ler arquivo Excel: {str(e)}")
                return redirect('core:user_import')

    else:
        form = UserImportForm()

    return render(request, 'core/user_import_form.html', {
        'form': form
    })



from django.shortcuts import render
from .forms import PublicVerificationForm
from .verification_engine import SecurityVerificationEngine

def public_verification(request):
    """
    Página pública para validação de autenticidade (Anti-fraude).
    Exibe os dados reais do documento se o hash for válido.
    """
    result = None
    
    if request.method == 'POST':
        form = PublicVerificationForm(request.POST)
        if form.is_valid():
            code = form.cleaned_data['verification_code']
            # O motor agora retorna um dicionário rico com dados seguros
            is_valid, data = SecurityVerificationEngine.process_query(code)
            
            result = {
                'valid': is_valid,
                'data': data, # Dicionário com metadados (ex: Nome, Curso, Ano, Status)
                'code': code
            }
    else:
        form = PublicVerificationForm()

    return render(request, 'public/verification.html', {
        'form': form,
        'result': result
    })


# Se este funcionar, apague a views acima (def public_verification(request):)
def public_student_search(request):
    """
    Página pública para confirmação de situação do aluno via ID.
    (Requisito 6)
    """
    student = None
    error_message = None

    if 'q' in request.GET:
        query_id = request.GET.get('q', '').strip()
        if query_id:
            try:
                # Pesquisa exata pelo registration_number (ID Automático)
                student = Student.objects.get(registration_number=query_id)
            except Student.DoesNotExist:
                error_message = "Nenhum aluno encontrado com este Número de Processo."
        else:
            error_message = "Por favor, insira um número de processo válido."

    return render(request, 'public/student_status_check.html', {
        'student': student,
        'error_message': error_message
    })


# apps/fiscal/views.py
from apps.core.utils import get_client_ip

@login_required
@transaction.atomic
def anular_documento_fiscal(request, doc_id):
    """
    Anula uma fatura confirmada e gera evidência de auditoria.
    Rigor: Apenas ADMIN ou Dr. Financeiro.
    """
    doc = get_object_or_404(DocumentoFiscal, id=doc_id, status='confirmed')
    
    # Segurança: Apenas cargos de alta confiança anulam faturas
    if request.user.current_role not in ['ADMIN', 'DIRECT_FINANC']:
        return HttpResponseForbidden("Operação restrita à Direção Financeira.")

    if request.method == 'POST':
        justificativa = request.POST.get('justificativa')
        
        if not justificativa or len(justificativa) < 15:
            messages.error(request, "ERRO: Justificativa muito curta ou ausente. Seja explícito.")
            return redirect('fiscal:documento_detail', doc_id=doc.id)

        # 1. Grava Auditoria Forense
        DocumentoCanceladoAudit.objects.create(
            documento=doc,
            operador=request.user,
            justificativa=justificativa,
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', 'Desconhecido'),
            valor_estornado=doc.valor_total
        )

        # 2. Altera Status Fiscal
        doc.status = 'cancelled'
        doc.save()

        # 3. DISPARO DE ALERTA (Simulação WhatsApp/Email para o Chefe)
        alert_msg = (
            f"⚠️ *ALERTA DE ANULAÇÃO FISCAL*\n"
            f"Documento: {doc.numero_documento}\n"
            f"Valor: {doc.valor_total:,.2f} Kz\n"
            f"Operador: {request.user.get_full_name()}\n"
            f"Motivo: {justificativa}\n"
            f"IP: {get_client_ip(request)}"
        )
        # task_send_urgent_alert.delay(settings.ADMIN_PHONE, alert_msg)

        messages.warning(request, f"Documento {doc.numero_documento} anulado. Ação reportada à auditoria.")
        return redirect('fiscal:saft_list')

    return render(request, 'fiscal/confirm_annulment.html', {'doc': doc})


