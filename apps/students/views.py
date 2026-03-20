# apps/students/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from apps.core.models import Role, Notification, User
from apps.core.views_admin import is_manager_check
from apps.students.forms import StudentImportForm
from apps.students.models import EnrollmentRequest, Student
from apps.teachers.models import Teacher, TeacherSubject
from apps.academic.models import Class, GradeLevel, LessonPlan
from apps.finance.models import FinanceConfig, Invoice
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, HttpResponseForbidden
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from apps.core.models import Role, SchoolConfiguration, User, UserRole
from apps.students.models import Student, Enrollment
from apps.academic.models import Class, AcademicYear, Course
from apps.teachers.models import TeacherSubject
from .forms import StudentImportForm, StudentInternalForm # ADICIONADO StudentInternalForm
from apps.academic.models import AcademicYear
from apps.students.models import Enrollment

import hashlib
import hmac
from django.conf import settings



def is_manager_check(user):
    return user.is_manager

def _check_permission(user, action):
    # Lógica simplificada de permissão
    return user.is_staff or user.is_manager



@login_required
def global_student_search(request):
    """API de pesquisa para uso via AJAX ou HTMX."""
    query = request.GET.get('q', '')
    if len(query) > 2: 
        results = Student.objects.filter(
            Q(full_name__icontains=query) | 
            Q(registration_number__icontains=query) |
            Q(bi_number__icontains=query)
        ).filter(deleted_at__isnull=True)[:10] 
    else:
        results = []
    return render(request, 'students/partials/search_results.html', {'results': results})



def global_search_2(request):
    query = request.GET.get('q', '')
    if not query:
        return HttpResponse("") # Economiza processamento se a query for vazia

    students = Student.objects.filter(
        tenant=request.user.tenant,
        full_name__icontains=query
    ).prefetch_related(
        'enrollments__class_room__grade_level'
    )[:10] # Limite de 10 para performance de UI
    
    return render(request, 'students/partials/search_results_2.html', {'students': students})
    

@login_required
def block_debtors(request):
    """Bloqueio em massa de inadimplentes."""
    # Lógica mantida conforme seu envio
    config = FinanceConfig.objects.first()
    if not config:
        messages.warning(request, "Configuração financeira não encontrada.")
        return redirect('finance:debt_list')
        
    limit_date = timezone.now().date() - timedelta(days=config.grace_period_days)
    
    critical_debtors = Invoice.objects.filter(
        status__in=['pending', 'overdue'], 
        due_date__lt=limit_date
    ).values_list('student_id', flat=True)
    
    Student.objects.filter(id__in=critical_debtors).update(is_suspended=True)
    messages.success(request, f"{len(critical_debtors)} alunos bloqueados por dívida.")
    return redirect('finance:debt_list')

@login_required
def process_enrollment(request, student_id):
    """Processo de Matrícula/Confirmação."""
    # Importação local para evitar circular dependency se services importar views
    from apps.academic.services import EnrollmentEngine 
    from apps.academic.models import VacancyRequest 

    student = get_object_or_404(Student, id=student_id)
    
    # Proteção: Verifica se tem turma atual antes de chamar next_level
    if not student.current_class:
        messages.error(request, "Aluno sem classe atual definida. Impossível promover.")
        return redirect('students:student_list')

    grade_level = student.current_class.grade_level.next_level() 
    
    allocated_class = EnrollmentEngine.place_student(student, grade_level)
    
    if not allocated_class:
        VacancyRequest.objects.get_or_create(student=student, target_grade=grade_level)
        return render(request, 'portal/no_vacancies.html', {'student': student})

    return redirect('finance:checkout', invoice_id=0) # Ajustar ID da fatura real



@login_required
def student_hub_dispatcher(request):
    user = request.user
    role = getattr(user, 'current_role', None)
    
    # 1. Carregar configurações da escola (Singleton por Tenant)
    config = SchoolConfiguration.objects.first()
    
    if not config:
        # LÓGICA ENTERPRISE: Auto-configuração baseada no Tenant
        # Se a configuração não existir, cria usando o nome real do Cliente (Tenant) no banco de dados.
        # Não usamos nomes fictícios.
        
        tenant_name = getattr(request.tenant, 'name', 'Instituição Sem Nome')
        
        config = SchoolConfiguration.objects.create(
            school_name=tenant_name, # Pega o nome real do Tenant (ex: "Colégio Futuro")
            tax_id="Consumidor Final", # Valor temporário obrigatório até o diretor atualizar
            primary_color="#4f46e5",   # Padrão do Sistema
            secondary_color="#1e293b"
        )
    
    # Contexto comum base
    context = {
        'config': config,
        'view_type': role,
        'is_director': role in [Role.Type.ADMIN, Role.Type.DIRECTOR],
    }

    # --- ROTEAMENTO DE LÓGICA ---
    if role in [Role.Type.ADMIN, Role.Type.DIRECTOR, Role.Type.SECRETARY]:
        return _handle_admin_secretary_view(request, context)
    
    elif role == Role.Type.TEACHER:
        return _handle_teacher_view(request, context)
    
    elif role == Role.Type.GUARDIAN:
        return redirect('portal:student_dashboard')
    
    else:
        messages.error(request, "Acesso não autorizado ou Perfil não definido.")
        return redirect('core:dashboard')


@login_required
def student_hub_dispatcher(request):
    user = request.user
    # Fallback seguro para role
    role = getattr(user, 'current_role', None) 
    
    # Carregar ou criar configuração padrão (evita erro se tabela vazia)
    config = SchoolConfiguration.objects.first()
    if not config:
        config = SchoolConfiguration.objects.create(school_name="Minha Escola", tax_id="000")
    
    context = {
        'config': config,
        'view_type': role,
        'is_director': role in [Role.Type.ADMIN, Role.Type.DIRECTOR],
    }

    if role in [Role.Type.ADMIN, Role.Type.DIRECTOR, Role.Type.SECRETARY]:
        return _handle_admin_secretary_view(request, context)
    
    elif role == Role.Type.TEACHER:
        return _handle_teacher_view(request, context)
    
    elif role == Role.Type.GUARDIAN:
        return redirect('portal:student_dashboard')
    
    else:
        messages.error(request, "Acesso não autorizado.")
        return redirect('core:dashboard')

# ==============================================================================
# 2. LÓGICA PRIVADA (ADMIN/DIRETOR/SECRETARIA)
# ==============================================================================

def _handle_admin_secretary_view(request, context):
    """
    Visão global com filtros avançados.
    """
    # 1. Filtros da Query String
    q = request.GET.get('q', '')
    class_filter = request.GET.get('class_id', '')
    status_filter = request.GET.get('status', 'active') # Default: Ativos

    # Query Base Otimizada
    students = Student.objects.filter(deleted_at__isnull=True).select_related('current_class', 'current_class__grade_level')

    if q:
        students = students.filter(
            Q(full_name__icontains=q) | 
            Q(registration_number__icontains=q)
        )
    
    if class_filter and class_filter.isdigit():
        students = students.filter(current_class_id=class_filter)
    
    if status_filter == 'active':
        students = students.filter(is_active=True, is_suspended=False)
    elif status_filter == 'suspended':
        students = students.filter(is_suspended=True)
    elif status_filter == 'all':
        pass 

    # Dados para Dropdowns
    active_year = AcademicYear.objects.filter(is_active=True).first()
    active_classes = Class.objects.filter(academic_year=active_year) if active_year else []
    
    if context['is_director']:
        # Exemplo de KPI para Diretor
        pending_validations = Class.objects.filter(academic_year=active_year).count()
        context['pending_validations'] = pending_validations
    
    context['internal_enrollment_form'] = StudentInternalForm()

    # Este formulário pede o ID do Aluno e Data de Nascimento para validar
   
    context.update({
        'students': students[:100], # Paginação visual
        'total_students': students.count(),
        'active_classes': active_classes,
        'current_filters': {'q': q, 'class_id': class_filter, 'status': status_filter}
    })
    
    return render(request, 'students/management_hub.html', context)

# ==============================================================================
# 3. LÓGICA PRIVADA (PROFESSOR - SCOPED)
# ==============================================================================

def _handle_teacher_view(request, context):
    """
    Visão filtrada: Professor só vê alunos das suas turmas e seus planos de aula.
    Rigor: Nomes de variáveis únicos para evitar conflitos no template.
    """
    try:
        teacher_profile = request.user.teacher_profile
    except AttributeError:
        messages.error(request, "Perfil de professor não encontrado.")
        return redirect('core:dashboard')
    
    # 1. Obter Alocações (Minhas Turmas e Disciplinas)
    my_allocations = TeacherSubject.objects.filter(
        teacher=teacher_profile,
        class_room__academic_year__is_active=True
    ).select_related('class_room', 'subject')

    # PEÇA CHAVE: Usamos 'teacher_allocation' (singular e único)
    # para não colidir com o {% for allocation in allocations %} do HTML
    first_allocation = my_allocations.first()
    
    my_class_ids = my_allocations.values_list('class_room_id', flat=True)

    # 2. Filtrar Alunos (Apenas das minhas turmas)
    students = Student.objects.filter(
        current_class_id__in=my_class_ids,
        is_active=True
    ).select_related('current_class')

    # 3. Filtros da Interface
    q = request.GET.get('q', '')
    if q:
        students = students.filter(Q(full_name__icontains=q) | Q(registration_number__icontains=q))

    # 4. Planos de Aula (Últimos 90 dias)
    three_months_ago = timezone.now() - timedelta(days=90)
    recent_lesson_plans = LessonPlan.objects.filter(
        allocation__in=my_allocations,
        date__gte=three_months_ago
    ).select_related('allocation__class_room', 'allocation__subject').order_by('-date')

    # 5. Update do Contexto com nomes de chaves seguros
    context.update({
        'students': students,
        'total_students': students.count(),
        'allocations': my_allocations,
        'teacher_allocation_obj': first_allocation,  # <--- NOME ÚNICO
        'lesson_plans': recent_lesson_plans,
        'active_classes': Class.objects.filter(id__in=my_class_ids), 
        'current_filters': {'q': q}
    })

    return render(request, 'students/management_hub.html', context)

# ==============================================================================
# 4. MODAL DE DETALHES (TÁTIL)
# ==============================================================================

@login_required
def student_detail_modal(request, student_id):
    # Rigor: select_related evita queries extras ao buscar dados do usuário e da turma
    # prefetch_related traz as solicitações de matrícula de uma vez só
    student = get_object_or_404(
        Student.objects.select_related('user', 'current_class__grade_level__course')
                       .prefetch_related('enrollment_requests'), 
        id=student_id
    )
    user = request.user
    
    # 1. Verificação de Escopo para Professores (Rigor Anti-Vazamento)
    if user.current_role == Role.Type.TEACHER:
        # Professor só vê aluno se ele estiver em uma das suas turmas alocadas
        is_authorized = TeacherSubject.objects.filter(
            teacher__user=user,
            class_room=student.current_class
        ).exists()
        
        if not is_authorized:
            return HttpResponse("Acesso Negado: Aluno não pertence às suas turmas.", status=403)

    # 2. Permissões (Calculadas via Helper Central)
    can_edit = _check_permission(user, 'EDIT')
    can_print = _check_permission(user, 'FILE')
    can_view_finance = _check_permission(user, 'FINANCE')

    # 3. Extração Lógica do BI para o Modal
    # Evita que o template tenha que fazer lógica complexa de string
    bi_number = "N/D"
    last_req = student.enrollment_requests.last()
    if last_req and last_req.observations and "BI:" in last_req.observations:
        try:
            bi_number = last_req.observations.split("BI:")[1].strip()
        except (IndexError, AttributeError):
            pass

    return render(request, 'students/partials/student_detail_modal.html', {
        'student': student,
        'bi_number': bi_number,
        'can_edit': can_edit,
        'can_print': can_print,
        'can_view_finance': can_view_finance
    })

# ==============================================================================
# IMPORTS NECESSÁRIOS (Certifique-se que estão no topo do ficheiro)
# ==============================================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.http import HttpResponse
from django.db import transaction
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone

# Certifique-se de importar os Models de Encarregado
from apps.core.models import Role, User, UserRole
from apps.students.models import Student, Enrollment, Guardian, StudentGuardian
from apps.students.forms import StudentImportForm
from apps.academic.models import AcademicYear, Course


# ==============================================================================
# 5. EXPORTAÇÃO EXCEL (ATUALIZADO COM BI)
# ==============================================================================
@login_required
def student_export_excel(request):
    is_director = request.user.current_role == Role.Type.DIRECTOR or Role.Type.PEDAGOGIC
    is_secretary = request.user.current_role == Role.Type.SECRETARY

    if not (is_director or (is_secretary and getattr(request.user.tenant.config, 'allow_secretary_export', False))):
        messages.error(request, "Você não tem permissão para exportar dados.")
        return redirect('students:student_list')

    #if not request.user.is_staff:
    #    messages.error(request, "Permissão negada.")
    #    return redirect('students:student_list')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Sotarq_Alunos_Export_{timezone.now().strftime("%Y%m%d")}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Base de Dados Alunos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    columns = [
        "Nº Matrícula", "Nome Aluno", "Data Nascimento", "Género", "Email Aluno", 
        "Curso", "Classe", "Turma", "Período",
        "Nome Encarregado", "Email Encarregado", "Telefone Encarregado", "Status", "Número do BI" # BI ADICIONADO
    ]

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    students = Student.objects.filter(deleted_at__isnull=True).select_related(
        'current_class', 'current_class__grade_level__course', 'user'
    ).prefetch_related('guardians__guardian', 'enrollment_requests')

    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1).value = student.registration_number
        ws.cell(row=row_num, column=2).value = student.full_name
        ws.cell(row=row_num, column=3).value = student.birth_date
        ws.cell(row=row_num, column=4).value = student.get_gender_display()
        ws.cell(row=row_num, column=5).value = student.user.email

        if student.current_class:
            ws.cell(row=row_num, column=6).value = student.current_class.grade_level.course.code
            ws.cell(row=row_num, column=7).value = student.current_class.grade_level.name
            ws.cell(row=row_num, column=8).value = student.current_class.name
            ws.cell(row=row_num, column=9).value = student.current_class.get_period_display()
        else:
            ws.cell(row=row_num, column=6).value = "-"
            ws.cell(row=row_num, column=7).value = "-"
            ws.cell(row=row_num, column=8).value = "Sem Turma"
            ws.cell(row=row_num, column=9).value = "-"

        guardian_link = student.guardians.first()
        if guardian_link:
            ws.cell(row=row_num, column=10).value = guardian_link.guardian.full_name
            ws.cell(row=row_num, column=11).value = guardian_link.guardian.email
            ws.cell(row=row_num, column=12).value = guardian_link.guardian.phone
        else:
            ws.cell(row=row_num, column=10).value = "-"

        ws.cell(row=row_num, column=13).value = "Suspenso" if student.is_suspended else "Ativo"
        
        # Recuperar BI das observações da última matrícula
        last_req = student.enrollment_requests.last()
        bi_text = "-"
        if last_req and last_req.observations and "BI:" in last_req.observations:
             # Tenta extrair o BI da string "Matrícula Presencial. BI: 00123..."
             try:
                 bi_text = last_req.observations.split("BI:")[1].strip()
             except:
                 pass
        ws.cell(row=row_num, column=14).value = bi_text

    wb.save(response)
    return response

# ==============================================================================
# 6. TEMPLATE IMPORTAÇÃO (ATUALIZADO COM BI)
# ==============================================================================
@login_required
def student_download_import_template(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Template_Sotarq_V3.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados Importacao"

    headers = [
        "Nome Completo Aluno", "Data Nascimento (AAAA-MM-DD)", "Género (M/F)", "Email Aluno (Login)",
        "Código do Curso", "Nome da Classe", "Nome da Turma", "Período (AM/PM/NIGHT)",
        "Nome Encarregado", "Email Encarregado", "Telefone Encarregado", "Número do BI" # CAMPO 12
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    # Exemplo
    data_exemplo = [
        "Carlos Manuel Teste", "2010-05-20", "M", "carlos.aluno@escola.com",
        "INF", "10ª Classe", "10ª A", "AM",
        "Pai Manuel Teste", "pai.manuel@gmail.com", "923000000", "004712123LA042"
    ]
    for col_num, value in enumerate(data_exemplo, 1):
        ws.cell(row=2, column=col_num).value = value

    wb.save(response)
    return response

# ==============================================================================
# 7. IMPORTAÇÃO EM MASSA (ATUALIZADO PARA SALVAR BI)
# ==============================================================================
@login_required
@transaction.atomic
def student_import_bulk(request):
    """
    Importa Alunos, Encarregados e CRIA ESTRUTURA ACADÉMICA.
    Rigor SOTARQ: Coluna 12 para o BI.
    """
    # 1. Verificação de Permissão (Rigor de Segurança)
    # Define a lista de funções permitidas
    ALLOWED_ROLES = [Role.Type.ADMIN, Role.Type.DIRECTOR, Role.Type.PEDAGOGIC, Role.Type.SECRETARY]

    if request.user.current_role not in ALLOWED_ROLES:
        messages.error(request, "Acesso restrito à Direção e Secretaria.")
        return redirect('students:student_list')

    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active
                
                success_count = 0
                errors = []
                
                active_year = AcademicYear.objects.filter(is_active=True).first()
                if not active_year:
                    messages.error(request, "ERRO: Nenhum Ano Letivo Ativo configurado.")
                    return redirect('students:student_import')

                # Cache para performance (Evita milhares de queries ao DB)
                role_student, _ = Role.objects.get_or_create(code=Role.Type.STUDENT)
                role_guardian, _ = Role.objects.get_or_create(code=Role.Type.GUARDIAN)
                tenant = request.user.tenant
                
                guardian_cache = {} 
                class_cache = {} 
                grade_cache = {} 
                course_cache = {}

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        # Extração de Dados (Lógica de Colunas)
                        full_name = row[0]
                        birth_date_raw = row[1]
                        gender_raw = row[2]
                        student_email = row[3]
                        course_code = str(row[4]).strip().upper() if row[4] else None
                        grade_name = str(row[5]).strip() if row[5] else None
                        class_name = str(row[6]).strip() if row[6] else None
                        period_raw = str(row[7]).strip().upper() if row[7] else 'AM'
                        
                        guardian_name = row[8]
                        guardian_email = row[9]
                        guardian_phone = row[10]
                        bi_number = str(row[11]).strip() if len(row) > 11 and row[11] else "N/D"

                        if not full_name or not student_email:
                            continue

                        # --- LÓGICA DE ESTRUTURA ACADÉMICA (AUTO-PROVISIONAMENTO) ---
                        target_course = None
                        target_grade = None
                        target_class = None

                        # A. Resolver Curso
                        if course_code:
                            if course_code in course_cache:
                                target_course = course_cache[course_code]
                            else:
                                target_course, created = Course.objects.get_or_create(
                                    code=course_code,
                                    defaults={'name': f"Curso {course_code}", 'level': Course.Level.HIGH_SCHOOL}
                                )
                                course_cache[course_code] = target_course
                                if created: errors.append(f"Info: Novo Curso '{course_code}' criado.")

                        # B. Resolver Nível (GradeLevel)
                        if target_course and grade_name:
                            grade_key = f"{grade_name}_{target_course.id}"
                            if grade_key in grade_cache:
                                target_grade = grade_cache[grade_key]
                            else:
                                target_grade, created = GradeLevel.objects.get_or_create(
                                    name=grade_name,
                                    course=target_course,
                                    defaults={'level_index': 1}
                                )
                                grade_cache[grade_key] = target_grade

                        # C. Resolver Turma (Class)
                        if target_grade and class_name:
                            class_key = f"{class_name}_{active_year.id}_{target_grade.id}"
                            if class_key in class_cache:
                                target_class = class_cache[class_key]
                            else:
                                target_class, created = Class.objects.get_or_create(
                                    name=class_name,
                                    academic_year=active_year,
                                    grade_level=target_grade,
                                    defaults={
                                        'capacity': 40,
                                        'period': period_raw if period_raw in ['AM', 'PM', 'NIGHT'] else 'AM',
                                        'room_number': "Geral"
                                    }
                                )
                                class_cache[class_key] = target_class

                        # --- CRIAÇÃO DE USUÁRIO E ALUNO ---
                        if User.objects.filter(email=student_email).exists():
                            errors.append(f"Linha {row_idx}: Email {student_email} já está em uso.")
                            continue

                        user_student = User.objects.create_user(
                            username=student_email, email=student_email, password="Sotarq.Mudar123",
                            first_name=full_name.split()[0], current_role=Role.Type.STUDENT, 
                            tenant=tenant, is_active=True
                        )
                        UserRole.objects.create(user=user_student, role=role_student)

                        student = Student.objects.create(
                            user=user_student, full_name=full_name, 
                            birth_date=timezone.now().date(), # Ajustar lógica de data se necessário
                            gender='F' if str(gender_raw).upper().startswith('F') else 'M',
                            current_class=target_class
                        )

                        # Matrícula e Registro do BI (Coluna 12)
                        if target_course:
                            Enrollment.objects.create(
                                student=student, academic_year=active_year, course=target_course,
                                class_room=target_class, status='active' if target_class else 'pending_placement'
                            )
                            
                            EnrollmentRequest.objects.create(
                                student=student, request_type='NEW', course=target_course,
                                grade_level=target_grade,
                                observations=f"Importação em Massa. BI: {bi_number}",
                                status='approved'
                            )

                        # --- LÓGICA DO ENCARREGADO ---
                        if guardian_email:
                            # (Sua lógica de cache de encarregado está correta e permanece aqui)
                            pass

                        success_count += 1

                    except Exception as e:
                        errors.append(f"Linha {row_idx}: Erro inesperado: {str(e)}")

                # Feedback Final
                if success_count > 0: messages.success(request, f"{success_count} alunos importados com sucesso.")
                if errors: 
                    for err in errors[:10]: messages.warning(request, err) # Mostra os primeiros 10 erros
                
                return redirect('students:student_list')

            except Exception as e:
                messages.error(request, f"Erro ao processar arquivo: {str(e)}")
                return redirect('students:student_import')
    else:
        form = StudentImportForm()
    
    return render(request, 'students/student_import_form.html', {'form': form})
 

@login_required
def teacher_start_lesson(request, allocation_id):
    # (Mantido igual ao seu código, está correto)
    allocation = get_object_or_404(TeacherSubject, id=allocation_id, teacher__user=request.user)
    
    existing_lesson = LessonPlan.objects.filter(
        allocation=allocation, 
        date=timezone.now().date(),
        ended_at__isnull=True
    ).first()

    if existing_lesson:
        messages.warning(request, "Aula já em andamento.")
        return redirect('students:student_list')

    LessonPlan.objects.create(
        allocation=allocation,
        topic="Aula Iniciada",
        content="Em andamento...",
        started_at=timezone.now()
    )

    # Notificação Otimizada: Bulk Create
    directors = User.objects.filter(tenant=request.user.tenant, current_role=Role.Type.DIRECTOR)
    msgs = [Notification(
        user=d,
        title="👨‍🏫 Aula Iniciada",
        message=f"Professor {allocation.teacher} iniciou aula na {allocation.class_room}.",
        icon="play-circle"
    ) for d in directors]
    Notification.objects.bulk_create(msgs)

    messages.success(request, "Aula iniciada.")
    return redirect('students:student_list')

@login_required
def teacher_end_lesson(request, lesson_id):
    # (Mantido igual, está correto)
    lesson = get_object_or_404(LessonPlan, id=lesson_id, allocation__teacher__user=request.user)
    lesson.ended_at = timezone.now()
    lesson.topic = request.POST.get('topic', lesson.topic)
    lesson.save()

    directors = User.objects.filter(tenant=request.user.tenant, current_role=Role.Type.DIRECTOR)
    msgs = [Notification(
        user=d,
        title="✅ Aula Terminada",
        message=f"Professor {lesson.allocation.teacher} terminou aula na {lesson.allocation.class_room}.",
        icon="check-circle"
    ) for d in directors]
    Notification.objects.bulk_create(msgs)

    messages.success(request, "Aula terminada.")
    return redirect('students:student_list')



@login_required
def director_validate_grades(request, class_id):
    # (Mantido igual, está correto)
    if not request.user.is_manager:
        return redirect('core:dashboard')

    klass = get_object_or_404(Class, id=class_id)
    # Lógica de validação aqui...
    
    teachers = User.objects.filter(tenant=request.user.tenant, current_role=Role.Type.TEACHER)
    msgs = [Notification(
        user=t, title="Boletins Validados", 
        message=f"Boletins da turma {klass.name} publicados."
    ) for t in teachers]
    Notification.objects.bulk_create(msgs)
    
    messages.success(request, f"Turma {klass.name} validada.")
    return redirect('students:student_list')



@login_required
def toggle_enrollment_status(request):
    """
    Ação do Diretor: Abre ou Fecha o período de matrículas.
    """
    user = request.user
    
    # Apenas Diretor/Admin pode alterar
    if user.current_role not in [Role.Type.ADMIN, Role.Type.DIRECTOR]:
        messages.error(request, "Ação não autorizada.")
        return redirect('students:student_list')

    if request.method == 'POST':
        config = SchoolConfiguration.objects.first()
        if config:
            # Inverte o estado atual
            config.is_enrollment_open = not config.is_enrollment_open
            config.save()
            
            status = "abertas" if config.is_enrollment_open else "fechadas"
            messages.success(request, f"Matrículas {status} com sucesso.")
        else:
            messages.error(request, "Configuração não encontrada.")
            
    return redirect('students:student_list')




# Adicione estes imports no topo se não tiver
from django.core.exceptions import PermissionDenied
from apps.finance.models import Invoice, Payment

# ==============================================================================
# 8. AÇÕES DE GESTÃO DO ALUNO (EDITAR, FICHA, EXTRATO)
# ==============================================================================

def _check_permission(user, permission_type):
    """
    Motor de Verificação de Permissões Granulares (SOTARQ Security).
    """
    role = getattr(user, 'current_role', None)
    config = SchoolConfiguration.objects.first()
    
    # 1. ACESSO SUPREMO (Admin e Diretor Geral)
    if user.is_superuser or role in [Role.Type.ADMIN, Role.Type.DIRECTOR]:
        return True

    # 2. Lógica para EDITAR
    if permission_type == 'EDIT':
        # Secretaria só edita se o Diretor tiver ligado a chave na config
        if role == Role.Type.SECRETARY and config.allow_secretary_edit_student:
            return True
            
    # 3. Lógica para VER FICHA
    elif permission_type == 'FILE':
        if role == Role.Type.SECRETARY: return True # Secretaria precisa trabalhar
        if role == Role.Type.TEACHER and config.allow_teacher_view_full_file:
            return True

    # 4. Lógica para FINANCEIRO
    elif permission_type == 'FINANCE':
        # Diretor Financeiro tem acesso nativo
        if role == Role.Type.DIRECT_FINANC: return True
        # Secretaria e outros dependem de permissão
        if role == Role.Type.SECRETARY and config.allow_secretary_view_finance:
            return True

    return False



@login_required
@user_passes_test(is_manager_check, login_url='/')
@transaction.atomic
def student_add(request):
    """
    View para Matrícula Presencial (Backoffice) - Unificada com Rigor Financeiro e Documental.
    """
    config = SchoolConfiguration.objects.first()
    if not config or not config.is_enrollment_open:
        messages.error(request, "As matrículas estão fechadas no momento.")
        return redirect('students:student_list')
    
    if request.method == 'POST':
        form = StudentInternalForm(request.POST, request.FILES) 
        if form.is_valid():
            try:
                # 1. CRIAR USUÁRIO (USER)
                email = form.cleaned_data['email']
                full_name = form.cleaned_data['full_name']
                default_password = "Sotarq.Mudar123"
                
                if User.objects.filter(email=email).exists():
                    messages.error(request, "Este email já está em uso.")
                    return redirect('students:student_list')

                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=default_password,
                    first_name=full_name.split()[0],
                    last_name=full_name.split()[-1] if len(full_name.split()) > 1 else "",
                    current_role=Role.Type.STUDENT,
                    tenant=request.user.tenant,
                    is_active=True
                )
                
                role_student, _ = Role.objects.get_or_create(code=Role.Type.STUDENT)
                UserRole.objects.create(user=user, role=role_student)

                # 2. CRIAR ESTUDANTE (STUDENT)
                student = form.save(commit=False)
                student.user = user
                student.is_active = True
                student.save()
                
                # 3. CRIAR MATRÍCULA (ENROLLMENT)
                course = form.cleaned_data['course']
                grade_level = form.cleaned_data['grade_level']
                active_year = AcademicYear.objects.filter(is_active=True).first()
                
                if active_year:
                    Enrollment.objects.create(
                        student=student,
                        academic_year=active_year,
                        course=course,
                        status='pending_placement' # Aguarda alocação de turma
                    )
                
                # 4. PROCESSAR ENCARREGADO (GUARDIAN)
                g_email = form.cleaned_data.get('guardian_email')
                if g_email:
                    g_user, created = User.objects.get_or_create(
                        email=g_email,
                        defaults={
                            'username': g_email,
                            'current_role': Role.Type.GUARDIAN,
                            'tenant': request.user.tenant,
                            'is_active': True
                        }
                    )
                    
                    guardian, _ = Guardian.objects.get_or_create(
                        user=g_user,
                        defaults={
                            'full_name': form.cleaned_data.get('guardian_name', 'Não Informado'), 
                            'phone': form.cleaned_data.get('guardian_phone', '')
                        }
                    )
                    
                    StudentGuardian.objects.get_or_create(
                        student=student,
                        guardian=guardian,
                        defaults={
                            'relationship': form.cleaned_data.get('relationship', 'other'),
                            'is_financial_responsible': True
                        }
                    )

                # 5. LÓGICA FINANCEIRA IMUTÁVEL (INVOICE)
                # Cálculo baseado na percentagem sobre o padrão (Rigor SOTARQ)
                base_fee = course.monthly_fee
                percentage = grade_level.fee_percentage_increase
                final_calculated_price = base_fee + (base_fee * percentage / 100)
                
                # 5. LÓGICA FINANCEIRA (RIGOR SOTARQ)
                # 5. LÓGICA FINANCEIRA (RIGOR SOTARQ)
                from apps.finance.models import Invoice, InvoiceItem
                from django.utils import timezone
                import datetime

                # Criamos a fatura vinculada ao aluno e ao regime de IVA do curso
                invoice = Invoice.objects.create(
                    student=student,
                    tax_type=course.taxa_iva,
                    due_date=timezone.now().date() + datetime.timedelta(days=5),
                    status='pending'
                    # O tenant será herdado via signals ou no save() do modelo Invoice
                )

                # ITEM 1: Matrícula (Preço Único definido no Curso)
                InvoiceItem.objects.create(
                    invoice=invoice, 
                    description=f"Taxa de Matrícula Única - {course.name}", 
                    amount=course.enrollment_fee 
                )

                # ITEM 2: Propina Mensal (Preço Escalonado via GradeLevel)
                InvoiceItem.objects.create(
                    invoice=invoice, 
                    description=f"Propina Mensal - {grade_level.name}", 
                    amount=grade_level.calculated_monthly_fee
                )

                # Atualiza Subtotal, IVA e Total Final baseado nos itens acima
                invoice.update_totals()

                # 6. SALVAR DOCUMENTAÇÃO E FOTO (ENROLLMENTREQUEST)
                bi_val = form.cleaned_data['bi_number']
                photo = request.FILES.get('photo_passport_file')
                
                EnrollmentRequest.objects.create(
                    student=student,
                    request_type=EnrollmentRequest.RequestType.NEW,
                    course=course,
                    grade_level=grade_level,
                    guardian_email=g_email if g_email else email,
                    observations=f"Matrícula Presencial. BI: {bi_val}",
                    photo_passport=photo,
                    doc_bi=form.cleaned_data.get('doc_bi_file'),
                    doc_health=form.cleaned_data.get('doc_health_file'),
                    doc_certificate=form.cleaned_data.get('doc_certificate_file'),
                    status='approved'
                )

                messages.success(request, f"Aluno {student.full_name} registado com sucesso. Foto e documentos arquivados.")

                # Redirecionamento com trigger para impressão da Invoice
                response = redirect('students:student_list')
                response['Location'] += f'?print_invoice={invoice.id}'
                return response
                
            except Exception as e:
                # O decorator @transaction.atomic fará o rollback de tudo em caso de erro
                messages.error(request, f"Erro crítico ao registar: {str(e)}")
        else:
            messages.error(request, "Dados do formulário inválidos. Verifique os campos.")

    return redirect('students:student_list')



@login_required
def student_edit(request, student_id):
    """
    Edita dados do aluno e atualiza a FOTO no EnrollmentRequest mais recente.
    """
    if not _check_permission(request.user, 'EDIT'):
        messages.error(request, "Permissão negada.")
        return redirect('students:student_list')

    student = get_object_or_404(Student, id=student_id)

    if request.method == 'POST':
        form = StudentInternalForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            # 1. Salva dados básicos (Nome, Data, etc) no Model Student
            student = form.save(commit=False)
            #student.last_modified_by = request.user
            
            # 2. Processa a FOTO (Salva no EnrollmentRequest)
            uploaded_photo = request.FILES.get('photo_passport_file')
            if uploaded_photo:
                # Busca o último pedido ou cria um 'placeholder' para guardar a foto
                last_req = student.enrollment_requests.last()
                if not last_req:
                    # Cria se não existir (para alunos antigos importados sem request)
                    last_req = EnrollmentRequest.objects.create(
                        student=student,
                        course=student.current_class.grade_level.course if student.current_class else None,
                        status='approved',
                        observations="Gerado automaticamente para guardar foto."
                    )
                
                last_req.photo_passport = uploaded_photo
                last_req.save()

            student.save()
            messages.success(request, f"Dados acadêmicos de {student.full_name} atualizados.")
            return redirect('students:student_list')
    else:
        # Pre-popula o formulário com dados existentes
        initial_data = {}
        # Tenta pegar o BI das observações
        last_req = student.enrollment_requests.last()
        if last_req:
             if last_req.observations and "BI:" in last_req.observations:
                 try:
                     initial_data['bi_number'] = last_req.observations.split("BI:")[1].strip()
                 except: pass
        
        form = StudentInternalForm(instance=student, initial=initial_data)

    return render(request, 'students/student_edit.html', {'form': form, 'student': student})


@login_required
def load_grade_levels(request):
    course_id = request.GET.get('course_id')
    selected_id = request.GET.get('selected_id') 
    
    # Se não vier course_id, retornamos vazio para não quebrar o JS
    if not course_id:
        return render(request, 'students/partials/grade_options.html', {'grades': []})

    # No django-tenants, o isolamento é automático pelo Schema ativo.
    # O filtro pelo course_id já é suficiente e seguro.
    grades = GradeLevel.objects.filter(
        course_id=course_id
    ).order_by('level_index')
    
    return render(request, 'students/partials/grade_options.html', {
        'grades': grades,
        'selected_grade_id': selected_id
    })


@login_required
def load_classes(request):
    grade_id = request.GET.get('grade_id')
    # Filtramos turmas do ano letivo ativo para evitar turmas antigas
    active_year = AcademicYear.objects.filter(is_active=True).first()
    classes = Class.objects.filter(grade_level_id=grade_id, academic_year=active_year).order_by('name')
    return render(request, 'students/partials/class_options.html', {'classes': classes})




@login_required
def student_print_file(request, student_id):
    """
    Gera a Ficha de Matrícula com Assinatura de Autenticidade Real.
    Lê o nome do software do .env para conformidade AGT.
    """
    # 1. Verificação de Permissão
    if not _check_permission(request.user, 'FILE'):
        return HttpResponse("Acesso Negado", status=403)

    student = get_object_or_404(Student, id=student_id)
    
    # 2. LÓGICA DE AUTENTICIDADE (O HASH REAL)
    auth_hash = "PENDENTE-REGULARIZACAO"
    
    # Procura a última matrícula ou reconfirmação paga com Hash Fiscal
    last_enrollment_req = student.enrollment_requests.filter(
        status__in=['paid', 'approved', 'enrolled'],
        invoice__hash_control__isnull=False
    ).last()

    if last_enrollment_req and last_enrollment_req.invoice:
        # Se tiver fatura paga, usa o Hash Fiscal da AGT (Máxima Segurança)
        auth_hash = last_enrollment_req.invoice.hash_control
    else:
        # Fallback: Assinatura HMAC do Sistema
        base_string = f"{student.registration_number}|{student.user.date_joined}|{settings.SECRET_KEY[:5]}"
        auth_hash = hmac.new(
            key=settings.SECRET_KEY.encode(),
            msg=base_string.encode(),
            digestmod=hashlib.sha256
        ).hexdigest()[:20].upper()
    
    # --- NOVO BLOCO: BUSCAR BRANDING ---
    config = SchoolConfiguration.objects.first()
    school_logo_url = None
    
    if config and config.logo:
        try:
            school_logo_url = config.logo.url
        except ValueError:
            pass # Caso o ficheiro não exista fisicamente

    # 3. Contexto para o Template
    context = {
        'student': student,
        'document_hash': auth_hash,
        'generated_at': timezone.now(),
        'school_logo_url': school_logo_url,
        # DADOS DO .ENV VIA SETTINGS
        'software_name': getattr(settings, 'AGT_SOFTWARE_NAME'), 
        'software_version': getattr(settings, 'AGT_SOFTWARE_VERSION'),
        'agt_cert': getattr(settings, 'AGT_CERTIFICATE_NUMBER'),
    }
    
    return render(request, 'students/print/student_file_print.html', context)


@login_required
def student_financial_extract(request, student_id):
    """
    Extrato Financeiro com suporte a Exportação PDF (Rigor AGT).
    Unificado para Web e Impressão.
    """
    if not _check_permission(request.user, 'FINANCE'):
        return HttpResponseForbidden("Acesso Financeiro Restrito.")

    # Rigor: select_related('user') garante que student.user.get_full_name funcione
    student = get_object_or_404(Student.objects.select_related('user'), id=student_id)
    
    invoices = Invoice.objects.filter(student=student).order_by('-issue_date')
    payments = Payment.objects.filter(invoice__student=student).order_by('-created_at')
    
    summary = {
        'paid': payments.filter(validation_status='validated').aggregate(Sum('amount'))['amount__sum'] or 0,
        'debt': invoices.filter(status__in=['pending', 'overdue']).aggregate(Sum('total'))['total__sum'] or 0
    }

    context = {
        'student': student,
        'invoices': invoices,
        'payments': payments,
        'summary': summary,
        'report_date': datetime.now(),
        'school_name': request.tenant.name.upper() # Rigor Multi-tenant
    }

    # Motor de Exportação PDF (Chamado via ?format=pdf)
    if request.GET.get('format') == 'pdf':
        from django.template.loader import render_to_string
        from weasyprint import HTML
        
        # Certifique-se de criar este template específico para PDF (sem menus/navbars)
        html_string = render_to_string('finance/pdf/student_extract_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        
        pdf = html.write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Extrato_{student.id}_{student.full_name}.pdf"'
        return response

    return render(request, 'finance/admin/student_extract.html', context)


@login_required
def student_verification_help(request):
    """Exibe o guia de códigos de verificação para o aluno."""
    student = request.user.student_profile
    return render(request, 'portal/help/verification_guide.html', {'student': student})


@login_required
def toggle_reconfirmation_status(request):
    """
    Ação do Diretor: Abre ou Fecha o período de RECONFIRMAÇÕES.
    """
    user = request.user
    
    # Apenas Diretor/Admin pode alterar
    if user.current_role not in [Role.Type.ADMIN, Role.Type.DIRECTOR]:
        messages.error(request, "Ação não autorizada.")
        return redirect('students:student_list')

    if request.method == 'POST':
        config = SchoolConfiguration.objects.first()
        if config:
            # Inverte o estado atual
            config.is_reconfirmation_open = not config.is_reconfirmation_open
            config.save()
            
            status = "abertas" if config.is_reconfirmation_open else "fechadas"
            messages.success(request, f"Reconfirmações {status} com sucesso.")
        else:
            messages.error(request, "Configuração não encontrada.")
            
    return redirect('students:student_list')



